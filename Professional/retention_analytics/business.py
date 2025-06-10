import os
import pandas  # Requirement
import zipfile
import mimetypes
import openpyxl  # Requirement
from datetime import datetime
import warnings
import re
warnings.simplefilter('ignore', category=UserWarning)

MBOB_POPULATED = False
REPORTS_POPULATED = False
CS_POPULATED = False
CRM_POPULATED = False

def swap_addresses(row):
    if pandas.notna(row['memberAddress2']) and not row['memberAddress1'].strip().startswith(tuple('0123456789')):
        return row['memberAddress2'], row['memberAddress1']
    return row['memberAddress1'], row['memberAddress2']

def swap_addresses_humana(row, col1, col2):
    if isinstance(row[col2], str) and row[col2].strip():
        return row[col2], row[col1]
    return row[col1], row[col2]

def format_phone(phone):
    if isinstance(phone, str) and len(phone) == 10 and phone.isdigit():
        return f"{phone[:3]}-{phone[3:6]}-{phone[6:]}"
    return phone

def parse_cigna_product(product):
    if not isinstance(product, str):
        return None, None
    parts = product.split('_')
    if len(parts) < 5:
        return None, None
    try:
        contract = parts[2]
        pbp = parts[3]
        plan_name = '_'.join(parts[4:])
        return f"{contract}-{pbp}", plan_name
    except:
        return None, None

def parse_aetna_name(name):
    if not isinstance(name, str) or not name.strip():
        return None, None
    parts = name.strip().split()
    if len(parts) < 2:
        return parts[0] if parts else None, None
    return parts[0], parts[-1]

def is_valid_excel(file_path):
    try:
        with zipfile.ZipFile(file_path, 'r') as zf:
            return True
    except zipfile.BadZipFile:
        return False

def generate_mbob_snapshot(snapshots_folder, data_folder, reports):
    print('\nGenerating snapshot...')
    base_headers = [
        'first', 'last', 'dob', 'mbi', 'email', 'phone', 'secondary_phone', 'address_1', 'address_2',
        'city', 'state', 'postal', 'county', 'second_address_1', 'second_address_2', 'second_city',
        'second_postal', 'second_county', 'carrier_1', 'product_1', 'identifier_1', 'contract_1',
        'effective_1', 'term_1', 'carrier_2', 'product_2', 'identifier_2', 'contract_2', 'effective_2', 'term_2'
    ]
    uhc_required_columns = ['agentName', 'mbiNumber', 'memberFirstName', 'memberLastName', 'planName', 'memberNumber',
                           'contract', 'pbp', 'segmentId', 'policyEffectiveDate', 'policyTermDate',
                           'memberAddress1', 'memberAddress2', 'secondaryAddressMemberCounty']
    uhc_column_mapping = {
        'memberFirstName': 'first',
        'memberLastName': 'last',
        'dateOfBirth': 'dob',
        'mbiNumber': 'mbi',
        'memberEmail': 'email',
        'memberPhone': 'phone',
        'secondaryPhoneNum': 'secondary_phone',
        'memberAddress1': 'address_1',
        'memberAddress2': 'address_2',
        'memberCity': 'city',
        'memberState': 'state',
        'memberZip': 'postal',
        'memberCounty': 'county',
        'secondaryAddressLine1': 'second_address_1',
        'secondaryAddressLine2': 'second_address_2',
        'secondaryAddressMemberCity': 'second_city',
        'secondaryAddressMemberZip': 'second_postal',
        'secondaryAddressMemberCounty': 'second_county',
        'planName': 'product_1',
        'memberNumber': 'identifier_1',
        'new_contract': 'contract_1',
        'policyEffectiveDate': 'effective_1',
        'policyTermDate': 'term_1'
    }
    humana_column_mapping = {
        'MbrLastName': 'last',
        'MbrFirstName': 'first',
        'Birth Date': 'dob',
        'Medicare No': 'mbi',
        'Email': 'email',
        'Primary Phone': 'phone',
        'Secondary Phone': 'secondary_phone',
        'Mail Address': 'second_address_1',
        'Mail Address 2': 'second_address_2',
        'Mail City': 'second_city',
        'Mail ZipCd': 'second_postal',
        'Mail Cnty': 'second_county',
        'Resident Address': 'address_1',
        'Resident Address 2': 'address_2',
        'Resident City': 'city',
        'Resident State': 'state',
        'Resident Zip Code': 'postal',
        'Resident County': 'county',
        'Plan Name': 'product_1',
        'Humana ID': 'identifier_1',
        'Contract-PBP-Segment ID': 'contract_1',
        'Effective Date': 'effective_1'
    }
    devoted_column_mapping = {
        'first_name': 'first',
        'last_name': 'last',
        'birth_date': 'dob',
        'email': 'email',
        'primary_phone': 'phone',
        'address': 'address_1',
        'address2': 'address_2',
        'city': 'city',
        'state': 'state',
        'zip_code': 'postal',
        'county': 'county',
        'pbp_name': 'product_1',
        'member_record_locator': 'identifier_1',
        'start_date': 'effective_1',
        'end_date': 'term_1'
    }
    cigna_column_mapping = {
        'First Name': 'first',
        'Last Name': 'last',
        'Date of Birth': 'dob',
        'Medicare Number': 'mbi',
        'Phone Number': 'phone',
        'Residential Address': 'address_1',
        'Residential Address Line 2': 'address_2',
        'Residential City': 'city',
        'Residential State': 'state',
        'Residential Zip Code': 'postal',
        'new_product': 'product_1',
        'Member ID': 'identifier_1',
        'new_contract': 'contract_1',
        'Effective Date': 'effective_1'
    }
    today = datetime.now().strftime('%Y-%m-%d')
    snapshot_path = os.path.join(snapshots_folder, f'Snapshot_{today}.xlsx')
    data_path = os.path.join(data_folder, f'{today}.xlsx')
    print('Processing UHC Book of Business...')
    uhc_bob = next((report for report in reports if report['name'] == 'UHC Book of Business'), None)
    uhc_df = pandas.read_excel(uhc_bob['location'], header=4)
    missing_uhc_columns = [col for col in uhc_required_columns if col not in uhc_df.columns]
    try:
        if missing_uhc_columns:
            raise KeyError(f"Missing required columns in UHC Book of Business: {missing_uhc_columns}\nPlease replace the report and reboot the software.\n")
    except KeyError as e:
        print(e)
        return
    if 'dateOfBirth' in uhc_df.columns:
        uhc_df['dateOfBirth'] = pandas.to_datetime(uhc_df['dateOfBirth'], errors='coerce').dt.strftime('%m/%d/%Y')
    uhc_df = uhc_df[uhc_df['agentName'] == 'INTEGRITY INSURANCE LLC']
    if 'mbiNumber' in uhc_df.columns:
        mbi_idx = uhc_df.columns.get_loc('mbiNumber')
        member_number_cols = [col for col in uhc_df.columns[:mbi_idx] if col == 'memberNumber']
        if len(member_number_cols) >= 2:
            rightmost_member_number = uhc_df.columns[mbi_idx - 1]
            if rightmost_member_number == 'memberNumber':
                uhc_df = uhc_df.drop(columns=rightmost_member_number)
    uhc_df['new_contract'] = uhc_df.apply(
        lambda row: f"{row['contract']}-{row['pbp']}-{row['segmentId']}", axis=1
    )
    uhc_df = uhc_df.drop(columns=['contract', 'pbp', 'segmentId'])
    uhc_df[['memberAddress1', 'memberAddress2']] = uhc_df.apply(
        swap_addresses, axis=1, result_type='expand'
    )
    uhc_df['secondaryAddressMemberCounty'] = uhc_df['secondaryAddressMemberCounty'].replace('null', '')
    mbob_df = pandas.DataFrame(columns=base_headers)
    for uhc_col, mbob_col in uhc_column_mapping.items():
        if uhc_col in uhc_df.columns:
            mbob_df[mbob_col] = uhc_df[uhc_col]
    mbob_df.loc[mbob_df['first'].notna() & mbob_df['last'].notna(), 'carrier_1'] = 'UHC'
    print('Processing Humana Book of Business...')
    humana_bob = next((report for report in reports if report['name'] == 'Humana Book of Business'), None)
    humana_df = pandas.read_excel(humana_bob['location'])
    humana_required_columns = ['Status', 'Humana ID', 'MbrFirstName', 'MbrLastName', 'Medicare No', 'Plan Name',
                              'Contract-PBP-Segment ID', 'Effective Date', 'Mail Address', 'Mail Address 2',
                              'Resident Address', 'Resident Address 2']
    missing_humana_columns = [col for col in humana_required_columns if col not in humana_df.columns]
    try:
        if missing_humana_columns:
            raise KeyError(f"Missing required columns in Humana Book of Business: {missing_humana_columns}\nPlease replace the report and reboot the software.\n")
    except KeyError as e:
        print(e)
        return
    if 'Birth Date' in humana_df.columns:
        humana_df['Birth Date'] = pandas.to_datetime(humana_df['Birth Date'], format='%m/%d/%Y', errors='coerce').dt.strftime('%m/%d/%Y')
    humana_df = humana_df[~humana_df['Status'].isin(['Inactive Policy', 'Cancelled Application'])]
    humana_df[['Mail Address', 'Mail Address 2']] = humana_df.apply(
        lambda row: swap_addresses_humana(row, 'Mail Address', 'Mail Address 2'), axis=1, result_type='expand'
    )
    humana_df[['Resident Address', 'Resident Address 2']] = humana_df.apply(
        lambda row: swap_addresses_humana(row, 'Resident Address', 'Resident Address 2'), axis=1, result_type='expand'
    )
    humana_df['Effective Date'] = pandas.to_datetime(humana_df['Effective Date'], format='%m/%d/%Y', errors='coerce').dt.strftime('%Y-%m-%d')
    humana_mbob_df = pandas.DataFrame(columns=base_headers)
    for humana_col, mbob_col in humana_column_mapping.items():
        if humana_col in humana_df.columns:
            humana_mbob_df[mbob_col] = humana_df[humana_col]
    humana_mbob_df.loc[humana_mbob_df['first'].notna() & humana_mbob_df['last'].notna(), 'carrier_1'] = 'Humana'
    print('Processing Devoted Book of Business...')
    devoted_bob = next((report for report in reports if report['name'] == 'Devoted Book of Business'), None)
    devoted_df = pandas.read_excel(devoted_bob['location'])
    devoted_required_columns = ['aor_policy_status', 'death_date', 'first_name', 'last_name', 'birth_date',
                               'email', 'primary_phone', 'address', 'address2', 'city', 'state', 'zip_code',
                               'county', 'pbp_name', 'member_record_locator', 'start_date', 'end_date']
    missing_devoted_columns = [col for col in devoted_required_columns if col not in devoted_df.columns]
    try:
        if missing_devoted_columns:
            raise KeyError(f"Missing required columns in Devoted Book of Business: {missing_devoted_columns}\nPlease replace the report and reboot the software.\n")
    except KeyError as e:
        print(e)
        return
    devoted_df = devoted_df[devoted_df['aor_policy_status'] != 'INACTIVE_POLICY']
    devoted_df = devoted_df[devoted_df['death_date'].isna()]
    if 'birth_date' in devoted_df.columns:
        devoted_df['birth_date'] = pandas.to_datetime(devoted_df['birth_date'], format='%Y-%m-%d', errors='coerce').dt.strftime('%m/%d/%Y')
    if 'start_date' in devoted_df.columns:
        devoted_df['start_date'] = pandas.to_datetime(devoted_df['start_date'], format='%Y-%m-%d', errors='coerce').dt.strftime('%Y/%m/%d')
    if 'end_date' in devoted_df.columns:
        devoted_df['end_date'] = pandas.to_datetime(devoted_df['end_date'], format='%Y-%m-%d', errors='coerce').dt.strftime('%Y/%m/%d')
    if 'primary_phone' in devoted_df.columns:
        devoted_df['primary_phone'] = devoted_df['primary_phone'].apply(format_phone)
    devoted_df['full_name'] = devoted_df['first_name'].str.strip().str.lower() + '|' + devoted_df['last_name'].str.strip().str.lower()
    mrl_counts = devoted_df['member_record_locator'].value_counts()
    for mrl in mrl_counts[mrl_counts > 1].index:
        dupe_rows = devoted_df[devoted_df['member_record_locator'] == mrl]
        main_row_idx = dupe_rows.index[0]
        main_row = devoted_df.loc[main_row_idx].copy()
        policy_idx = 1
        for idx in dupe_rows.index[1:]:
            policy_idx += 1
            main_row[f'pbp_name_{policy_idx}'] = dupe_rows.loc[idx, 'pbp_name']
            main_row[f'member_record_locator_{policy_idx}'] = dupe_rows.loc[idx, 'member_record_locator']
            main_row[f'start_date_{policy_idx}'] = dupe_rows.loc[idx, 'start_date']
            main_row[f'end_date_{policy_idx}'] = dupe_rows.loc[idx, 'end_date']
        devoted_df.loc[main_row_idx] = main_row
        devoted_df = devoted_df.drop(dupe_rows.index[1:])
    devoted_df = devoted_df.drop(columns=['full_name'])
    devoted_mbob_df = pandas.DataFrame(columns=base_headers)
    for devoted_col, mbob_col in devoted_column_mapping.items():
        if devoted_col in devoted_df.columns:
            devoted_mbob_df[mbob_col] = devoted_df[devoted_col]
    devoted_mbob_df.loc[devoted_mbob_df['first'].notna() & devoted_mbob_df['last'].notna(), 'carrier_1'] = 'Devoted'
    print('Processing Cigna Book of Business...')
    cigna_bob = next((report for report in reports if report['name'] == 'Cigna Book of Business'), None)
    cigna_df = pandas.read_excel(cigna_bob['location'])
    cigna_required_columns = ['First Name', 'Last Name', 'Date of Birth', 'Medicare Number', 'Phone Number',
                             'Residential Address', 'Residential Address Line 2', 'Residential City',
                             'Residential State', 'Residential Zip Code', 'Product', 'Member ID', 'Effective Date']
    missing_cigna_columns = [col for col in cigna_required_columns if col not in cigna_df.columns]
    try:
        if missing_cigna_columns:
            raise KeyError(f"Missing required columns in Cigna Book of Business: {missing_cigna_columns}\nPlease replace the report and reboot the software.\n")
    except KeyError as e:
        print(e)
        return
    cigna_df['full_name'] = cigna_df['First Name'].str.strip().str.lower() + '|' + cigna_df['Last Name'].str.strip().str.lower()
    medicare_counts = cigna_df['Medicare Number'].value_counts()
    for medicare in medicare_counts[medicare_counts > 1].index:
        dupe_rows = cigna_df[cigna_df['Medicare Number'] == medicare]
        main_row_idx = dupe_rows.index[0]
        main_row = cigna_df.loc[main_row_idx].copy()
        policy_idx = 1
        for idx in dupe_rows.index[1:]:
            policy_idx += 1
            main_row[f'Product_{policy_idx}'] = dupe_rows.loc[idx, 'Product']
            main_row[f'Member ID_{policy_idx}'] = dupe_rows.loc[idx, 'Member ID']
            main_row[f'Effective Date_{policy_idx}'] = dupe_rows.loc[idx, 'Effective Date']
        cigna_df.loc[main_row_idx] = main_row
        cigna_df = cigna_df.drop(dupe_rows.index[1:])
    cigna_df = cigna_df.drop(columns=['full_name'])
    if 'Date of Birth' in cigna_df.columns:
        cigna_df['Date of Birth'] = pandas.to_datetime(cigna_df['Date of Birth'], format='%m/%d/%Y', errors='coerce').dt.strftime('%m/%d/%Y')
    if 'Effective Date' in cigna_df.columns:
        cigna_df['Effective Date'] = pandas.to_datetime(cigna_df['Effective Date'], format='%m/%d/%Y', errors='coerce').dt.strftime('%Y/%m/%d')
    if 'Phone Number' in cigna_df.columns:
        cigna_df['Phone Number'] = cigna_df['Phone Number'].apply(format_phone)
    cigna_df[['new_contract', 'new_product']] = cigna_df['Product'].apply(parse_cigna_product).apply(pandas.Series)
    cigna_mbob_df = pandas.DataFrame(columns=base_headers)
    for cigna_col, mbob_col in cigna_column_mapping.items():
        if cigna_col in cigna_df.columns:
            cigna_mbob_df[mbob_col] = cigna_df[cigna_col]
    cigna_mbob_df.loc[cigna_mbob_df['first'].notna() & cigna_mbob_df['last'].notna(), 'carrier_1'] = 'Cigna'
    print('Processing Aetna Book of Business...')
    aetna_bob = next((report for report in reports if report['name'] == 'Aetna Book of Business'), None)
    aetna_df = pandas.read_excel(aetna_bob['location'])
    aetna_required_columns = ['Customer name', 'Customer no', 'Effective date']
    missing_aetna_columns = [col for col in aetna_required_columns if col not in aetna_df.columns]
    try:
        if missing_aetna_columns:
            raise KeyError(f"Missing required columns in Aetna Book of Business: {missing_aetna_columns}\nPlease replace the report and reboot the software.\n")
    except KeyError as e:
        print(e)
        return
    aetna_df[['First', 'Last']] = aetna_df['Customer name'].apply(parse_aetna_name).apply(pandas.Series)
    if 'Effective date' in aetna_df.columns:
        aetna_df['Effective date'] = pandas.to_datetime(aetna_df['Effective date'], format='%m/%d/%Y', errors='coerce').dt.strftime('%Y/%m/%d')
    aetna_column_mapping = {
        'First': 'first',
        'Last': 'last',
        'Customer no': 'identifier_1',
        'Effective date': 'effective_1'
    }
    aetna_mbob_df = pandas.DataFrame(columns=base_headers)
    for aetna_col, mbob_col in aetna_column_mapping.items():
        if aetna_col in aetna_df.columns:
            aetna_mbob_df[mbob_col] = aetna_df[aetna_col]
    aetna_mbob_df.loc[aetna_mbob_df['first'].notna() & aetna_mbob_df['last'].notna(), 'carrier_1'] = 'Aetna'
    aetna_mbob_df.loc[aetna_mbob_df['dob'].isna(), 'duplicate_flag'] = 'missing_dob'
    with warnings.catch_warnings():
        warnings.filterwarnings('ignore', category=FutureWarning)
        mbob_df = pandas.concat([mbob_df, humana_mbob_df, devoted_mbob_df, cigna_mbob_df, aetna_mbob_df], ignore_index=True)
    mbob_df['first'] = mbob_df['first'].str.strip().str.lower()
    mbob_df['last'] = mbob_df['last'].str.strip().str.lower()
    mbob_df['full_name'] = mbob_df['first'] + '|' + mbob_df['last']
    mbob_df['dob'] = mbob_df['dob'].astype(str).replace('NaT', '')
    headers = base_headers.copy()
    mbob_df['duplicate_flag'] = mbob_df['duplicate_flag'].fillna('')
    mbob_df['full_name_dob'] = mbob_df['full_name'] + '|' + mbob_df['dob']
    full_name_dob_counts = mbob_df[mbob_df['dob'] != '']['full_name_dob'].value_counts()
    max_policies = full_name_dob_counts.max() if not full_name_dob_counts.empty else 2
    if max_policies > 2:
        for i in range(3, max_policies + 1):
            headers.extend([f'carrier_{i}', f'product_{i}', f'identifier_{i}', f'contract_{i}', f'effective_{i}', f'term_{i}'])
    mbob_df = mbob_df.reindex(columns=headers + ['duplicate_flag', 'full_name', 'full_name_dob'])
    for fnd in full_name_dob_counts[full_name_dob_counts > 1].index:
        dupe_rows = mbob_df[mbob_df['full_name_dob'] == fnd]
        main_row_idx = dupe_rows.index[0]
        main_row = mbob_df.loc[main_row_idx].copy()
        main_row['first'] = max(dupe_rows['first'], key=len)
        carriers = dupe_rows['carrier_1'].unique()
        policy_idx = 1
        assigned_pdp = False
        rx_row = dupe_rows[dupe_rows['product_1'].str.contains('rx', case=False, na=False)]
        if not rx_row.empty and len(dupe_rows) >= 2:
            rx_row_idx = rx_row.index[0]
            main_row['product_2'] = dupe_rows.loc[rx_row_idx, 'product_1']
            main_row['identifier_2'] = dupe_rows.loc[rx_row_idx, 'identifier_1']
            main_row['contract_2'] = dupe_rows.loc[rx_row_idx, 'contract_1']
            main_row['effective_2'] = dupe_rows.loc[rx_row_idx, 'effective_1']
            main_row['term_2'] = dupe_rows.loc[rx_row_idx, 'term_1']
            main_row['carrier_2'] = dupe_rows.loc[rx_row_idx, 'carrier_1']
            assigned_pdp = True
            policy_idx = 2
        for idx in dupe_rows.index:
            if assigned_pdp and idx == rx_row.index[0]:
                continue
            target_idx = 2 if policy_idx == 2 and not assigned_pdp else policy_idx if policy_idx != 2 else 1
            main_row[f'product_{target_idx}'] = dupe_rows.loc[idx, 'product_1']
            main_row[f'identifier_{target_idx}'] = dupe_rows.loc[idx, 'identifier_1']
            main_row[f'contract_{target_idx}'] = dupe_rows.loc[idx, 'contract_1']
            main_row[f'effective_{target_idx}'] = dupe_rows.loc[idx, 'effective_1']
            main_row[f'term_{target_idx}'] = dupe_rows.loc[idx, 'term_1']
            main_row[f'carrier_{target_idx}'] = dupe_rows.loc[idx, 'carrier_1']
            policy_idx += 1 if policy_idx != 2 else 2
        mbob_df.loc[main_row_idx] = main_row
        for idx in dupe_rows.index:
            if idx != main_row_idx:
                mbob_df.loc[idx, ['product_1', 'identifier_1', 'contract_1', 'effective_1', 'term_1']] = pandas.NA
    mbob_df = mbob_df[(mbob_df['product_1'].notna()) | (mbob_df['carrier_1'] == 'Aetna')]
    mbob_df = mbob_df.drop(columns=['full_name', 'full_name_dob'])
    mbob_df.to_excel(snapshot_path, index=False)
    wb = openpyxl.load_workbook(snapshot_path)
    ws = wb.active
    duplicate_flag_col = mbob_df.columns.get_loc('duplicate_flag') + 1
    for row in range(2, ws.max_row + 1):
        flag = ws.cell(row=row, column=duplicate_flag_col).value
    wb.save(snapshot_path)
    wb.save(data_path)
    print(f'Success!\nSnapshot_{today}.xlsx can be found in the "Snapshots" folder.')

def consolidate_commission_statements():
    print('\033[91mNot available in sample.\n\033[0m')

def calculate_retention_rates(root_directory):
    data_folder = find_data_folder(root_directory)
    snapshot_files = []
    # Identify snapshot files by naming convention
    for file_name in os.listdir(data_folder):
        if file_name.startswith('Snapshot_') and file_name.endswith('.xlsx'):
            match = re.match(r'Snapshot_(\d{4}-\d{2}-\d{2})\.xlsx', file_name)
            if match:
                snapshot_files.append({
                    'name': file_name,
                    'location': os.path.join(data_folder, file_name),
                    'date': match.group(1)
                })
    
    if not snapshot_files:
        print("No snapshot files found in the Program Data folder.")
        return
    
    # Extract years and months
    months_data = {}
    for snapshot in snapshot_files:
        date_str = snapshot['date']
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            year = date_obj.strftime('%Y')
            month = date_obj.strftime('%m')
            mmyyyy = date_obj.strftime('%m%Y')
            
            if mmyyyy not in months_data:
                months_data[mmyyyy] = {
                    'Name': mmyyyy,
                    'Year': year,
                    'Month': month,
                    'Count': 0,
                    'Client Count': 0,
                    'Files': []
                }
            
            months_data[mmyyyy]['Count'] += 1
            months_data[mmyyyy]['Files'].append(snapshot['location'])
            
            # Read the spreadsheet to get row count (excluding header)
            df = pandas.read_excel(snapshot['location'], engine='openpyxl')
            months_data[mmyyyy]['Client Count'] += len(df)  # Excludes header by default
            
        except ValueError:
            continue
    
    # Calculate average Client Count for each MMYYYY
    for mmyyyy in months_data:
        count = months_data[mmyyyy]['Count']
        if count > 0:
            months_data[mmyyyy]['Client Count'] = round(months_data[mmyyyy]['Client Count'] / count, 2)
    
    # Create a list of unique years
    years = sorted(set(data['Year'] for data in months_data.values()))
    months = [f'{m:02d}' for m in range(1, 13)]
    
    # Initialize the retention rate DataFrame
    retention_data = pandas.DataFrame(index=months, columns=years + ['Annual Retention'])
    retention_data.index.name = 'Month'
    
    # Calculate monthly retention rates
    sorted_months = sorted(months_data.keys())
    for i, mmyyyy in enumerate(sorted_months):
        current_count = months_data[mmyyyy]['Client Count']
        # Find the last non-null month before the current one
        for prev_mmyyyy in sorted_months[:i][::-1]:
            prev_count = months_data[prev_mmyyyy]['Client Count']
            if prev_count > 0:
                retention_rate = (current_count / prev_count) * 100 if prev_count > 0 else 0
                year = months_data[mmyyyy]['Year']
                month = months_data[mmyyyy]['Month']
                retention_data.loc[month, year] = round(retention_rate, 2)
                break
    
    # Calculate annual retention rate
    if sorted_months:
        oldest_mmyyyy = sorted_months[0]
        newest_mmyyyy = sorted_months[-1]
        oldest_count = months_data[oldest_mmyyyy]['Client Count']
        newest_count = months_data[newest_mmyyyy]['Client Count']
        if oldest_count > 0:
            annual_retention = (newest_count / oldest_count) * 100
            retention_data['Annual Retention'] = round(annual_retention, 2)
    
    # Save the retention rates to a spreadsheet
    output_path = os.path.join(root_directory, 'Retention_Rates.xlsx')
    retention_data.to_excel(output_path, engine='openpyxl')
    print(f"Retention rates saved to {output_path}")

def define_data(data_folder):
    global MBOB_POPULATED
    data = []
    treasure = [
        'first', 'last', 'dob', 'mbi', 'email', 'phone', 'secondary_phone', 'address_1', 'address_2',
        'city', 'state', 'postal', 'county', 'second_address_1', 'second_address_2', 'second_city',
        'second_postal', 'second_county', 'carrier_1', 'product_1', 'identifier_1', 'contract_1',
        'effective_1', 'term_1'
    ]
    matching_files = 0
    for file_name in os.listdir(data_folder):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(data_folder, file_name)
            df = pandas.read_excel(file_path, nrows=1)
            headers = df.columns.tolist()
            headers_lower = [h.lower() for h in headers]
            treasure_lower = [t.lower() for t in treasure]
            if all(t in headers_lower for t in treasure_lower):
                matching_files += 1
                data.append({'name': file_name, 'location': file_path})
    MBOB_POPULATED = matching_files >= 2
    return data

def define_spreadsheets(reports_folder):
    global REPORTS_POPULATED
    global CS_POPULATED
    global CRM_POPULATED Nostalgic
    reports = []
    spreadsheets_found = []
    to_convert = []
    potential_hits = []
    missing_required = []
    matched_reports = set()
    valid_mime = [
        {'type': 'xls', 'mime': ['application/vnd.ms-excel', 'application/xls'], 'extension': '.xls'},
        {'type': 'xlsx', 'mime': ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'], 'extension': '.xlsx'},
        {'type': 'xlsm', 'mime': ['application/vnd.ms-excel.sheet.macroEnabled.12'], 'extension': '.xlsm'},
        {'type': 'xlsb', 'mime': ['application/vnd.ms-excel.sheet.binary.macroEnabled.12'], 'extension': '.xlsb'},
        {'type': 'csv', 'mime': ['text/csv'], 'extension': '.csv'},
        {'type': 'txt', 'mime': ['text/plain'], 'extension': '.txt'}
    ]
    treasure = [
        {'name': 'UHC Book of Business', 'headers': ['nmaName80', 'nmaWidStatus80', 'nmaNpn80', 'nmaName70'], 'found': False, 'required': True},
        {'name': 'Humana Book of Business', 'headers': ['Humana ID', 'Contract-PBP-Segment ID', 'Deeming Indicator', 'AOR SAN'], 'found': False, 'required': True},
        {'name': 'Devoted Book of Business', 'headers': ['__typename', 'nuggets.__typename', 'working_application_status', 'working_enrollment_id'], 'found': False, 'required': True},
        {'name': 'Cigna Book of Business', 'headers': ['AFMO', 'AFMO Writing ID', 'NFMO', 'NFMO Writing ID'], 'found': False, 'required': True},
        {'name': 'Aetna Book of Business', 'headers': ['Product 1', 'Product 2', 'Product 3', 'Product 4'], 'found': False, 'required': True},
        {'name': 'Radiusbob Report', 'headers': ['Part A Effective Date', 'Part B Effective Date', 'Veteran Status', 'Medicare ID Number'], 'found': False, 'required': False},
        {'name': 'UHC Commission Statement', 'headers': ['Client Reference #', 'AARP Member ID / Exchange ID', 'Area ID', 'Reduction Indicator'], 'found': False, 'required': False},
        {'name': 'Humana Commission Statement', 'headers': ['AorSan', 'AdvOrCbkPrm', 'TxnTypeCd', 'PriorPolNbr'], 'found': False, 'required': False},
        {'name': 'Devoted Commission Statement', 'headers': ['Description', 'Member Count', 'Total', 'Payee'], 'found': False, 'required': False}
    ]
    for file_name in os.listdir(reports_folder):
        file_path = os.path.join(reports_folder, file_name)
        if os.path.isfile(file_path):
            extension = os.path.splitext(file_name)[1].lower()
            if any(vm['extension'] == extension for vm in valid_mime):
                spreadsheets_found.append(file_path)
    for file_path in spreadsheets_found:
        extension = os.path.splitext(file_path)[1].lower()
        if extension == '.xlsx' and is_valid_excel(file_path):
            mime_type, _ = mimetypes.guess_type(file_path)
            valid_xlsx_mime = next(vm['mime'] for vm in valid_mime if vm['type'] == 'xlsx')
            if mime_type in valid_xlsx_mime:
                potential_hits.append(file_path)
            else:
                to_convert.append(file_path)
        else:
            to_convert.append(file_path)
    for file_path in to_convert:
        try:
            extension = os.path.splitext(file_path)[1].lower()
            new_path = os.path.splitext(file_path)[0] + '_converted.xlsx'
            if extension == '.csv':
                df = pandas.read_csv(file_path, low_memory=False)
            elif extension == '.txt':
                df = pandas.read_csv(file_path, sep='\t')
            else:
                df = pandas.read_excel(file_path)
            df.to_excel(new_path, index=False, engine='openpyxl')
            if is_valid_excel(new_path):
                potential_hits.append(new_path)
                os.remove(file_path)
            else:
                os.remove(new_path)
        except Exception as e:
            continue
    for file_path in potential_hits:
        try:
            if not is_valid_excel(file_path):
                continue
            df = pandas.read_excel(file_path, nrows=5, engine='openpyxl')
            headers = list(df.columns.str.lower())
            row_5 = list(df.iloc[4].astype(str).str.lower()) if len(df) >= 5 else []
            headers_uhc = []
            if not any(t['found'] for t in treasure if t['name'] == 'UHC Book of Business'):
                df_check = pandas.read_excel(file_path, nrows=0, engine='openpyxl')
                df_full = pandas.read_excel(file_path, engine='openpyxl')
                if len(df_full) >= 5:
                    df_uhc = pandas.read_excel(file_path, header=4, nrows=1, engine='openpyxl')
                    headers_uhc = list(df_uhc.columns.str.lower())
            for t in treasure:
                if t['name'] in matched_reports:
                    continue
                t_headers = [h.lower() for h in t['headers']]
                matched = False
                if t['name'] == 'UHC Book of Business' and headers_uhc:
                    matched = any(h in headers_uhc for h in t_headers)
                else:
                    matched = any(h in headers for h in t_headers) or any(h in row_5 for h in t_headers)
                if matched:
                    t['found'] = True
                    reports.append({'name': t['name'], 'location': file_path})
                    matched_reports.add(t['name'])
                    break
        except Exception as e:
            continue
    for t in treasure:
        if t['required'] and not t['found']:
            missing_required.append(t['name'])
    CRM_POPULATED = any(t['name'] == 'Radiusbob Report' and t['found'] for t in treasure)
    CS_POPULATED = all(t['found'] for t in treasure if not t['required'] and t['name'] != 'Radiusbob Report')
    if not missing_required:
        REPORTS_POPULATED = True
    if REPORTS_POPULATED == False or CRM_POPULATED == False or CS_POPULATED == False:
        print('\nMISSING FILES')
        if missing_required:
            for name in missing_required:
                print(f'{name}')
        if CRM_POPULATED == False or CS_POPULATED == False:
            for t in treasure:
                if t['required'] == False and t['found'] == False:
                    print(f'{t["name"]}')
        print('')
    return reports

def find_data_folder(root_directory):
    data_folder = os.path.join(root_directory, 'Program Data')
    if not os.path.exists(data_folder):
        print('Creating missing "Program Data" folder...')
        os.makedirs(data_folder)
    return data_folder

def find_snapshots_folder(root_directory):
    snapshots_folder = os.path.join(root_directory, 'Snapshots')
    if not os.path.exists(snapshots_folder):
        print('Creating missing "Snapshots" folder...')
        os.makedirs(snapshots_folder)
    return snapshots_folder

def find_reports_folder(root_directory):
    reports_folder = os.path.join(root_directory, 'Reports')
    if not os.path.exists(reports_folder):
        print('Creating missing "Reports" folder...')
        os.makedirs(reports_folder)
    return reports_folder