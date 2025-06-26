import os
import re
import uuid
import random
import shutil
import warnings
import webbrowser
import tkinter as tk
from datetime import datetime
from tkinter import messagebox, filedialog, ttk
import win32com.client as win32 # Requirement
import pandas as pd # Requirement
from pandas.errors import EmptyDataError # Requirement
from openpyxl import Workbook # Requirement
from openpyxl import load_workbook # Requirement
from scipy.stats import pearsonr # Requirement

# Ignore only the 'no default style' warning from openpyxl
warnings.filterwarnings(
    'ignore',
    message = "Workbook contains no default style, apply openpyxl's default"
)

# --- HELPERS ------------------------------------------------------------------------------------------------------

def find_data_directory(directories, permission_errors):
    try:
        data_entry = next(d for d in directories if d.get('name') == 'Program Data')
        cwd = os.getcwd()
        data_path = os.path.join(cwd, 'Program Data')

        if os.path.isdir(data_path):
            data_entry['found'] = True
            data_entry['location'] = data_path
        else:
            os.makedirs(data_path)
            data_entry['found'] = True
            data_entry['location'] = data_path

    except PermissionError:
        permission_errors['state'] = True
        return

def find_snapshots_directory(directories, permission_errors):
    try:
        snapshots_entry = next(d for d in directories if d.get('name') == 'Snapshots')
        cwd = os.getcwd()
        snapshots_path = os.path.join(cwd, 'Snapshots')

        if os.path.isdir(snapshots_path):
            snapshots_entry['found'] = True
            snapshots_entry['location'] = snapshots_path
        else:
            os.makedirs(snapshots_path)
            snapshots_entry['found'] = True
            snapshots_entry['location'] = snapshots_path

    except PermissionError:
        permission_errors['state'] = True
        return

def find_reports_directory(directories, valid_generation, permission_errors):
    try:
        reports_entry = next(d for d in directories if d['name'] == 'Reports')
        cwd = os.getcwd()
        reports_path = os.path.join(cwd, 'Reports')

        if os.path.isdir(reports_path):
            reports_entry['found'] = True
            reports_entry['location'] = reports_path
        else:
            os.makedirs(reports_path)
            reports_entry['found'] = True
            reports_entry['location'] = reports_path
            valid_generation['state'] = False

    except PermissionError:
        permission_errors['state'] = True

def convert_all_to_xlsx(reports_directory):
    folder = reports_directory['location']
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False

    for fname in os.listdir(folder):
        base, ext = os.path.splitext(fname)
        ext = ext.lower()
        if ext not in ('.xls'):
            continue

        old_path = os.path.join(folder, fname)
        new_path = os.path.join(folder, f"{base}.xlsx")

        try:
            # Open with Excel
            wb = excel.Workbooks.Open(old_path)
            # 51 = xlOpenXMLWorkbook (no macro)
            wb.SaveAs(new_path, FileFormat=51)
            wb.Close()
            os.remove(old_path)
        except Exception as e:
            os_errors['state'] = True
            return

    excel.Quit()

def find_uhc_data(reports_directory, carrier_reports, permission_errors, os_errors, switch):
    folder = reports_directory.get('location')
    uhc_entry = next((r for r in carrier_reports if r['name'] == 'UHC'), None)
    expected_headers = uhc_entry['headers']

    try:
        for fname in os.listdir(folder):
            path = os.path.join(folder, fname)
            _, ext = os.path.splitext(fname.lower())
            if ext in ('.csv', '.xls', '.xlsx'):
                
                # Read with row 5 as header
                try:
                    if switch == 'GS':
                        if ext == '.csv':
                            df = pd.read_csv(path, header=4, dtype=str)
                        else:
                            df = pd.read_excel(path, header=4, dtype=str)
                    elif switch == 'CS':
                        if ext == '.csv':
                            df = pd.read_csv(path, header=0, dtype=str)
                        else:
                            df = pd.read_excel(path, header=0, dtype=str)
                except PermissionError:
                    # Cannot open/read this file
                    permission_errors['state'] = True
                    return
                except (EmptyDataError, ValueError, IndexError):
                    continue

                # Check for exact header match
                cols = list(df.columns)
                if all(h in cols for h in expected_headers):
                    uhc_entry['found'] = True
                    uhc_entry['location'] = path
                    break

    except PermissionError:
        permission_errors['state'] = True
    except OSError:
        os_errors['state'] = True
    return

def find_humana_data(reports_directory, carrier_reports, permission_errors, os_errors):
    folder = reports_directory.get('location')
    humana_entry = next((r for r in carrier_reports if r['name'] == 'Humana'), None)
    expected_headers = humana_entry['headers']

    try:
        for fname in os.listdir(folder):
            path = os.path.join(folder, fname)
            _, ext = os.path.splitext(fname.lower())
            if ext in ('.csv', '.xls', '.xlsx'):
                try:
                    if ext == '.csv':
                        df = pd.read_csv(path, header=0, dtype=str)
                    else:
                        df = pd.read_excel(path, header=0, dtype=str)
                except PermissionError:
                    # Cannot open/read this file
                    permission_errors['state'] = True
                    return
                except (EmptyDataError, ValueError, IndexError):
                    continue

                # Check for exact header match
                cols = list(df.columns)
                if all(h in cols for h in expected_headers):
                    humana_entry['found'] = True
                    humana_entry['location'] = path
                    break

    except PermissionError:
        permission_errors['state'] = True
    except OSError:
        os_errors['state'] = True
    return

def find_aetna_data(reports_directory, carrier_reports, permission_errors, os_errors):
    folder = reports_directory.get('location')
    aetna_entry = next((r for r in carrier_reports if r['name'] == 'Aetna'), None)
    expected_headers = aetna_entry['headers']

    try:
        for fname in os.listdir(folder):
            path = os.path.join(folder, fname)
            _, ext = os.path.splitext(fname.lower())
            if ext in ('.csv', '.xls', '.xlsx'):
                try:
                    if ext == '.csv':
                        df = pd.read_csv(path, header=0, dtype=str)
                    else:
                        df = pd.read_excel(path, header=0, dtype=str)
                except PermissionError:
                    # Cannot open/read this file
                    permission_errors['state'] = True
                    return
                except (EmptyDataError, ValueError, IndexError):
                    continue

                # Check for exact header match
                cols = list(df.columns)
                if all(h in cols for h in expected_headers):
                    aetna_entry['found'] = True
                    aetna_entry['location'] = path
                    break

    except PermissionError:
        permission_errors['state'] = True
    except OSError:
        os_errors['state'] = True
    return

def create_dated_spreadsheet(snapshots_directory, permission_errors, os_errors, switch):
    try:
        folder = snapshots_directory.get('location')

        # build filename with date
        date_str = datetime.now().strftime('%Y%m%d')
        if switch == 'GS':
            filename = f"snapshot_{date_str}.xlsx"
        if switch == 'CS':
            filename = f"ccs_{date_str}.xlsx"
        file_path = os.path.join(folder, filename)

        # create workbook and write headers
        wb = Workbook()
        ws = wb.active
        if switch == 'GS':
            headers = [
                "first_name", "last_name", "mbi", "dob", "email", "phone",
                "address", "address_2", "city", "county", "state", "postal_code",
                "product_1", "identifier_1", "effective_date_1", "term_date_1", "carrier_1"
            ]
        if switch == 'CS':
            headers = [
                'name', 'product', 'paid', 'comment', 'YTD paid'
                ]
        ws.append(headers)

        # save the file
        wb.save(file_path)
        return file_path

    except PermissionError:
        # file system permission issue
        permission_errors['state'] = True
        return None

    except OSError:
        # any other OS-related error (e.g. disk full, invalid path)
        os_errors['state'] = True
        return None

def copy_snapshot_to_data_directory(snapshot_file, data_directory, os_errors):
    try:
        base_name = os.path.basename(snapshot_file)
        root, ext = os.path.splitext(base_name)
        dest_path = os.path.join(data_directory, root)
        shutil.copy2(snapshot_file, dest_path)
    except OSError:
        os_errors['state'] = True
        return

# --- HELP ---------------------------------------------------------------------------------------------------------

def display_help():
    print("This program accepts Book of Business spreadsheets from: UHC, Humana, Devoted, and Aetna.\n"
          + "This program also accepts Commission Statement spreadsheets from: UHC, Humana, and Aetna.\n"
          + "\nDownloaded Book of Business / Commission Statement spreadsheets should be placed in the 'Reports' folder.\n"
          + "Afterwards, you can click 'Generate Snapshot' or 'Generate CCS' to compile these spreadsheets into data points.\n"
          + "\nYou can find the Book of Business spreadsheets in the following locations:\n"
          + "UHC: Jarvis -> Book of Business tab -> Download results, (Default settings)\nHumana: AgentPortal -> View all Customers -> Export, (Default settings)\n"
          + "Aetna: ProducerWorld -> Customer & Compensation Summary -> Book of Business for: Integrity Insurance LLC -> Download as CSV\n"
          + "\nYou can find the Commission Statement spreadsheets in the following locations:\n"
          + "UHC: Jarvis -> Commissions, (Statements and More) -> Download\nHumana: AgentPortal -> Commission Statement Portal -> Excel\n"
          + "The Aetna Book of Business spreadsheet also doubles as the Commission Statement spreadsheet, and does not have to be downloaded twice.\n"
          + "\nOnce two or more data points have been generated, you can click the 'View Analytics' button to compare and view simple metrics.\n"
          + "Data points are saved in the 'Snapshots' folder as Excel files. Feel free to play with them - this program stores it's own backups!")

# --- COMPARE DATA -------------------------------------------------------------------------------------------------

def compare_data():
    permission_errors = {'state': False}
    directories = [
        {'name': 'Program Data', 'found': False, 'location': None}
    ]

    # Find Program Data directory
    find_data_directory(directories, permission_errors)
    if permission_errors['state']:
        messagebox.showerror("Permission Error", 
                            "Permission error encountered while accessing Program Data directory.\n"
                            "Please run the program with administrator privileges and try again.")
        return
    if not directories[0]['found']:
        messagebox.showerror("Directory Error", 
                            "Program Data directory not found.\n"
                            "Please ensure the directory exists and try again.")
        return

    data_dir = directories[0]['location']
    
    # Scan for snapshot files (extensionless, starting with 'snapshot_')
    snapshot_files = [f for f in os.listdir(data_dir) if f.startswith('snapshot_')]
    if len(snapshot_files) < 2:
        messagebox.showerror("Insufficient Files", 
                            "At least two snapshot files are required for comparison.\n"
                            "Please generate more snapshots and try again.")
        return

    # Create GUI for file selection
    root = tk.Tk()
    root.title("Select Snapshot Files for Comparison")
    root.geometry("600x600")
    root.minsize(400, 400)  # Set minimum window size

    tk.Label(root, text="Select two snapshot files to compare:", font=("Arial", 12)).pack(pady=10)

    # Listbox for file selection
    listbox = tk.Listbox(root, selectmode=tk.MULTIPLE, width=50, height=10)
    for file in snapshot_files:
        listbox.insert(tk.END, file)
    listbox.pack(pady=10, padx=10, fill=tk.X)

    # Scrollable text widget for results
    text_frame = tk.Frame(root)
    text_frame.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
    
    scrollbar = tk.Scrollbar(text_frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    result_text = tk.Text(text_frame, wrap=tk.WORD, height=15, yscrollcommand=scrollbar.set, font=("Arial", 10))
    result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.config(command=result_text.yview)
    
    result_text.config(state='disabled')  # Make text widget read-only initially

    def on_compare():
        selected = listbox.curselection()
        if len(selected) != 2:
            result_text.config(state='normal')
            result_text.delete(1.0, tk.END)
            result_text.insert(tk.END, "Please select exactly two files for comparison.\n")
            result_text.config(state='disabled')
            return

        file1 = snapshot_files[selected[0]]
        file2 = snapshot_files[selected[1]]
        temp_file1 = os.path.join(data_dir, f"{file1}.xlsx")
        temp_file2 = os.path.join(data_dir, f"{file2}.xlsx")

        try:
            # Temporarily rename files by appending '.xlsx'
            os.rename(os.path.join(data_dir, file1), temp_file1)
            os.rename(os.path.join(data_dir, file2), temp_file2)

            # Read the snapshot files
            df1 = pd.read_excel(temp_file1, dtype=str).fillna('')
            df2 = pd.read_excel(temp_file2, dtype=str).fillna('')

            # Create primary key for comparison
            df1['primary_key'] = df1['first_name'].str.strip() + df1['last_name'].str.strip() + df1['dob'].str.strip()
            df2['primary_key'] = df2['first_name'].str.strip() + df2['last_name'].str.strip() + df2['dob'].str.strip()

            # Identify added, removed, and common clients
            clients1 = set(df1['primary_key'])
            clients2 = set(df2['primary_key'])
            added = clients2 - clients1
            removed = clients1 - clients2
            common = clients1 & clients2

            differences = []

            # Report removed clients
            if removed:
                differences.append("Clients no longer on the book:")
                for pk in removed:
                    client = df1[df1['primary_key'] == pk][['first_name', 'last_name']].iloc[0]
                    differences.append(f"  {client['first_name']} {client['last_name']}")
                differences.append("")

            # Report added clients
            if added:
                differences.append("Clients added to the book:")
                for pk in added:
                    client = df2[df2['primary_key'] == pk][['first_name', 'last_name']].iloc[0]
                    differences.append(f"  {client['first_name']} {client['last_name']}")
                differences.append("")

            # Compare common clients for changes
            if common:
                differences.append("Changes in client details:")
                for pk in common:
                    row1 = df1[df1['primary_key'] == pk].iloc[0]
                    row2 = df2[df2['primary_key'] == pk].iloc[0]
                    client_name = f"{row1['first_name']} {row1['last_name']}"
                    changed = False
                    changes = [f"  {client_name}:"]
                    
                    for col in df1.columns:
                        if col == 'primary_key':
                            continue
                        val1 = str(row1[col]).strip()
                        val2 = str(row2[col]).strip()
                        if val1 != val2:
                            changes.append(f"    {col}: '{val1}' -> '{val2}'")
                            changed = True
                    
                    if changed:
                        differences.extend(changes)

            if not (added or removed or any("Changes in client details:" in d for d in differences)):
                differences.append("No differences found between the selected snapshots.")

            # Display results in text widget
            result_text.config(state='normal')
            result_text.delete(1.0, tk.END)
            result_text.insert(tk.END, "\n".join(differences) + "\n")
            result_text.config(state='disabled')

        except Exception as e:
            result_text.config(state='normal')
            result_text.delete(1.0, tk.END)
            result_text.insert(tk.END, f"Error during comparison: {str(e)}\n")
            result_text.config(state='disabled')
        
        finally:
            # Restore original filenames by removing '.xlsx'
            try:
                if os.path.exists(temp_file1):
                    os.rename(temp_file1, os.path.join(data_dir, file1))
                if os.path.exists(temp_file2):
                    os.rename(temp_file2, os.path.join(data_dir, file2))
            except Exception as e:
                result_text.config(state='normal')
                result_text.delete(1.0, tk.END)
                result_text.insert(tk.END, f"Error restoring original filenames: {str(e)}\n")
                result_text.config(state='disabled')

    tk.Button(root, text="Compare Selected Files", command=on_compare).pack(pady=10)
    tk.Button(root, text="Close", command=root.destroy).pack(pady=5)

    root.mainloop()

# --- VIEW ANALYTICS -----------------------------------------------------------------------------------------------

def view_analytics():
    permission_errors = {'state': False}
    
    # Use the directory of logic.py (Program Data) as the data directory
    data_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Scan for snapshot and CCS files in Program Data
    snapshot_files = [f for f in os.listdir(data_dir) if f.startswith('snapshot_')]
    ccs_files = [f for f in os.listdir(data_dir) if f.startswith('ccs_')]
    
    # Debug: Log detected files
    print(f"Detected snapshot files: {snapshot_files}")
    print(f"Detected CCS files: {ccs_files}")

    snapshot_valid = len(snapshot_files) >= 2
    ccs_valid = len(ccs_files) >= 1

    if not (snapshot_valid or ccs_valid):
        print("At least two snapshot files or one CCS file required for analytics.")
        return

    temp_files = []
    analytics_data = []
    try:
        # Temporarily append '.xlsx' to files
        for file in snapshot_files + ccs_files:
            original_path = os.path.join(data_dir, file)
            temp_path = os.path.join(data_dir, f"{file}.xlsx")
            try:
                os.rename(original_path, temp_path)
                temp_files.append((file, temp_path))
            except Exception as e:
                print(f"Error renaming {file}: {str(e)}")
                continue

        # Process snapshot files for client counts
        if snapshot_valid:
            for file, temp_path in [(f, t) for f, t in temp_files if f.startswith('snapshot_')]:
                try:
                    df = pd.read_excel(temp_path, dtype=str).fillna('')
                    match = re.search(r'(\d{8})', file)
                    if not match:
                        print(f"Skipping {file}: Invalid date format")
                        continue
                    date_str = match.group(1)
                    try:
                        file_date = datetime.strptime(date_str, '%Y%m%d').strftime('%Y-%m-%d')
                    except ValueError:
                        print(f"Skipping {file}: Invalid date {date_str}")
                        continue

                    client_count = len(df)
                    analytics_data.append({
                        'date': file_date,
                        'client_count': client_count,
                        'filename': file,
                        'commission_sum': None  # Use None for missing commission data
                    })
                except Exception as e:
                    print(f"Error reading snapshot {file}: {str(e)}")
                    continue

        # Process CCS files for commission sums
        commission_changes = []
        if ccs_valid:
            # Group CCS files by year
            ccs_by_year = {}
            for file, temp_path in [(f, t) for f, t in temp_files if f.startswith('ccs_')]:
                match = re.search(r'(\d{4})(\d{2})(\d{2})', file)
                if not match:
                    print(f"Skipping CCS {file}: Invalid date format")
                    continue
                year = match.group(1)
                if year not in ccs_by_year:
                    ccs_by_year[year] = []
                ccs_by_year[year].append((file, temp_path))

            # Process each year separately
            for year, files in ccs_by_year.items():
                # Find the latest CCS file for this year
                latest_ccs_file, latest_ccs_path = max(
                    files,
                    key=lambda f: re.search(r'(\d{8})', f[0]).group(1) if re.search(r'(\d{8})', f[0]) else '0',
                    default=(None, None)
                )
                latest_ccs_date = None
                if latest_ccs_file:
                    match = re.search(r'(\d{8})', latest_ccs_file)
                    latest_ccs_date = datetime.strptime(match.group(1), '%Y%m%d').strftime('%Y-%m-%d')

                # Calculate Aetna's YTD paid for this year
                aetna_ytd_sum = 0
                if latest_ccs_path:
                    try:
                        df_latest = pd.read_excel(latest_ccs_path, dtype=str).fillna('0')
                        if 'YTD paid' in df_latest.columns:
                            aetna_ytd_sum = pd.to_numeric(df_latest['YTD paid'], errors='coerce').fillna(0).sum()
                    except Exception as e:
                        print(f"Error reading latest CCS file {latest_ccs_file}: {str(e)}")

                # Distribute Aetna's YTD paid across CCS files in this year
                num_ccs_files = len(files)
                aetna_per_file = aetna_ytd_sum / num_ccs_files if num_ccs_files > 0 else 0

                # Process each CCS file in this year
                for file, temp_path in files:
                    try:
                        df = pd.read_excel(temp_path, dtype=str).fillna('0')
                        paid_sum = pd.to_numeric(df['paid'], errors='coerce').fillna(0).sum()
                        commission_sum = paid_sum + aetna_per_file

                        match = re.search(r'(\d{8})', file)
                        if not match:
                            print(f"Skipping CCS {file}: Invalid date format")
                            continue
                        date_str = match.group(1)
                        try:
                            file_date = datetime.strptime(date_str, '%Y%m%d').strftime('%Y-%m-%d')
                        except ValueError:
                            print(f"Skipping CCS {file}: Invalid date {date_str}")
                            continue

                        # Update or add analytics_data
                        for data in analytics_data:
                            if data['date'] == file_date:
                                data['commission_sum'] = round(commission_sum, 2)
                                break
                        else:
                            analytics_data.append({
                                'date': file_date,
                                'client_count': 0,
                                'filename': file,
                                'commission_sum': round(commission_sum, 2)
                            })
                    except Exception as e:
                        print(f"Error reading CCS file {file}: {str(e)}")
                        continue

        if len(analytics_data) < 2:
            print("Insufficient valid files for analytics.")
            return

        # Sort by date
        analytics_data.sort(key=lambda x: x['date'])

        # Calculate retention rates (only for snapshot files)
        retention_rates = []
        if snapshot_valid:
            snapshot_data = [d for d in analytics_data if d['filename'].startswith('snapshot_')]
            for i in range(1, len(snapshot_data)):
                prev_data = snapshot_data[i-1]
                curr_data = snapshot_data[i]
                prev_file = os.path.join(data_dir, f"{prev_data['filename']}.xlsx")
                curr_file = os.path.join(data_dir, f"{curr_data['filename']}.xlsx")
                try:
                    df_prev = pd.read_excel(prev_file, dtype=str).fillna('')
                    df_curr = pd.read_excel(curr_file, dtype=str).fillna('')
                    df_prev['primary_key'] = df_prev['first_name'].str.strip() + df_prev['last_name'].str.strip() + df_prev['dob'].str.strip()
                    df_curr['primary_key'] = df_curr['first_name'].str.strip() + df_curr['last_name'].str.strip() + df_curr['dob'].str.strip()
                    prev_clients = set(df_prev['primary_key'])
                    curr_clients = set(df_curr['primary_key'])
                    retained = len(prev_clients & curr_clients)
                    retention_rate = (retained / len(prev_clients)) * 100 if len(prev_clients) > 0 else 0
                    retention_rates.append({
                        'period': f"{prev_data['date']} to {curr_data['date']}",
                        'retention_rate': round(retention_rate, 2),
                        'month': prev_data['date'][:7],  # YYYY-MM for aggregation
                        'end_date': curr_data['date']  # For chart alignment
                    })
                except Exception as e:
                    print(f"Error comparing snapshot {prev_data['filename']} and {curr_data['filename']}: {str(e)}")
                    continue

        # Calculate commission changes
        commission_changes = []
        if ccs_valid:
            valid_comm_data = [d for d in analytics_data if d['commission_sum'] is not None]
            for i in range(1, len(valid_comm_data)):
                prev_data = valid_comm_data[i-1]
                curr_data = valid_comm_data[i]
                prev_comm = prev_data['commission_sum']
                curr_comm = curr_data['commission_sum']
                commission_change = ((curr_comm - prev_comm) / prev_comm * 100) if prev_comm > 0 else 0
                commission_changes.append({
                    'period': f"{prev_data['date']} to {curr_data['date']}",
                    'commission_change': round(commission_change, 2),
                    'month': prev_data['date'][:7],  # YYYY-MM for aggregation
                    'end_date': curr_data['date']  # For chart alignment
                })

        # Detect outliers in commission changes (z-scores)
        outliers = []
        if commission_changes:
            commission_values = [c['commission_change'] for c in commission_changes]
            mean_comm = sum(commission_values) / len(commission_values) if commission_values else 0
            std_comm = (sum((x - mean_comm) ** 2 for x in commission_values) / len(commission_values)) ** 0.5 if commission_values else 0
            for c in commission_changes:
                z_score = abs(c['commission_change'] - mean_comm) / std_comm if std_comm > 0 else 0
                if z_score > 2:  # Threshold for outliers
                    outliers.append({
                        'period': c['period'],
                        'commission_change': c['commission_change'],
                        'z_score': round(z_score, 2)
                    })

        if not (retention_rates or commission_changes):
            print("No retention rates or commission changes could be calculated.")
            return

        # Calculate Pearson correlation using paired periods
        pearson_corr = None
        pearson_message = ""
        if snapshot_valid and ccs_valid and retention_rates and commission_changes:
            # Pair retention rates and commission changes by closest end_date
            paired_data = []
            for r in retention_rates:
                r_date = r['end_date']
                # Find the closest commission change by end_date
                closest_comm = min(
                    commission_changes,
                    key=lambda c: abs((datetime.strptime(c['end_date'], '%Y-%m-%d') - 
                                      datetime.strptime(r_date, '%Y-%m-%d')).days),
                    default=None
                )
                if closest_comm:
                    paired_data.append((r['retention_rate'], closest_comm['commission_change']))
            
            if len(paired_data) >= 2:
                retention_values, commission_values = zip(*paired_data)
                # Check for constant values
                if len(set(retention_values)) == 1 or len(set(commission_values)) == 1:
                    pearson_message = "Correlation undefined due to constant values"
                else:
                    try:
                        pearson_corr, _ = pearsonr(retention_values, commission_values)
                    except Exception as e:
                        print(f"Error calculating Pearson correlation: {str(e)}")
                        pearson_message = f"Correlation calculation failed: {str(e)}"

        # Calculate averages
        avg_retention = round(sum(r['retention_rate'] for r in retention_rates) / len(retention_rates), 2) if retention_rates else 0
        avg_commission_change = round(sum(c['commission_change'] for c in commission_changes) / len(commission_changes), 2) if commission_changes else 0

        # Generate HTML output
        title = "Client Retention and Commission Analytics" if snapshot_valid and ccs_valid else \
                "Client Retention Analytics" if snapshot_valid else "Commission Analytics"
        html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@3.9.1/dist/chart.min.js"></script>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f4f4f4;
            display: flex;
            flex-direction: column;
            min-height: 100vh;
        }}
        h1, h2 {{
            color: #333;
            text-align: center;
        }}
        .table-container {{
            width: 90%;
            max-height: 300px;
            overflow-y: auto;
            margin: 20px auto;
            background-color: #fff;
            box-shadow: 0 0 10px rgba(0,0,0,0.1);
            flex-grow: 1;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border: 1px solid #ddd;
        }}
        th {{
            background-color: #4CAF50;
            color: white;
            position: sticky;
            top: 0;
            z-index: 1;
        }}
        .commission-th {{
            background-color: #800080;
            color: white;
        }}
        .outlier-th {{
            background-color: #FF4500;
            color: white;
        }}
        tr:nth-child(even) {{
            background-color: #f2f2f2;
        }}
        .chart-container {{
            flex-grow: 1;
            width: 98%;
            margin: 20px auto;
            display: flex;
            justify-content: center;
            align-items: stretch;
            min-height: 400px;
        }}
        canvas {{
            width: 100% !important;
            max-width: none;
            height: auto !important;
            margin: 20px auto;
            display: block;
        }}
        .summary {{
            text-align: center;
            font-size: 1.2em;
            margin: 20px 0;
        }}
    </style>
</head>
<body>
    <h1>{title}</h1>
    <div class="summary">
        {"<p>Average All-Time Retention Rate: {}%</p>".format(avg_retention) if snapshot_valid else ""}
        {"<p>Average All-Time Commission Change: {}%</p>".format(avg_commission_change) if ccs_valid else ""}
        {"<p>Pearson Correlation (Retention vs. Commission Change): {}</p>".format(round(pearson_corr, 2)) if pearson_corr is not None else f"<p>Pearson Correlation: {pearson_message}</p>"}
    </div>
    <div class="table-container">
        <table>
            <tr>
                <th>Date</th>
                {"<th>Client Count</th><th>Retention Rate (%)</th>" if snapshot_valid else ""}
                {"<th class='commission-th'>Commission Sum ($)</th><th class='commission-th'>Commission Change (%)</th>" if ccs_valid else ""}
            </tr>
"""
        # Match retention and commission changes by end_date
        retention_dict = {r['end_date']: r['retention_rate'] for r in retention_rates}
        commission_dict = {c['end_date']: c['commission_change'] for c in commission_changes}
        for data in analytics_data:
            curr_date = data['date']
            retention = retention_dict.get(curr_date, '-')
            commission_change = commission_dict.get(curr_date, '-')
            commission_sum = data['commission_sum'] if data['commission_sum'] is not None else 'N/A'
            html_content += f"""
            <tr>
                <td>{data['date']}</td>
                {"<td>{}</td><td>{}</td>".format(data['client_count'], retention) if snapshot_valid else ""}
                {"<td>{}</td><td>{}</td>".format(commission_sum, commission_change) if ccs_valid else ""}
            </tr>
"""
        html_content += """
        </table>
    </div>
"""
        # Add outliers table
        if outliers:
            html_content += """
    <h2>Suspicious Commission Fluctuations</h2>
    <div class="table-container">
        <table>
            <tr>
                <th class="outlier-th">Period</th>
                <th class="outlier-th">Commission Change (%)</th>
                <th class="outlier-th">Z-Score</th>
            </tr>
"""
            for outlier in outliers:
                html_content += f"""
            <tr>
                <td>{outlier['period']}</td>
                <td>{outlier['commission_change']}</td>
                <td>{outlier['z_score']}</td>
            </tr>
"""
            html_content += """
        </table>
    </div>
"""

        # Prepare chart data with single y-axis
        retention_labels = [r['end_date'] for r in retention_rates]
        commission_labels = [c['end_date'] for c in commission_changes]
        all_labels = sorted(set(retention_labels + commission_labels))
        
        retention_data = []
        for label in all_labels:
            retention_data.append(
                next((r['retention_rate'] for r in retention_rates if r['end_date'] == label), None)
            )
        
        commission_data = []
        for label in all_labels:
            commission_data.append(
                next((c['commission_change'] for c in commission_changes if c['end_date'] == label), None)
            )

        html_content += """
    <h2>Analytics Chart</h2>
    <div class="chart-container">
        <canvas id="analyticsChart"></canvas>
    </div>
    <script>
        const ctx = document.getElementById('analyticsChart').getContext('2d');
        const chart = new Chart(ctx, {
            type: 'line',
            data: {
                labels: ['""" + "', '".join(all_labels) + """'],
                datasets: [""" + \
                ("""
                    {
                        label: 'Retention Rate (%)',
                        data: [""" + ", ".join(str(x) if x is not None else 'null' for x in retention_data) + """],
                        borderColor: '#4CAF50',
                        backgroundColor: 'rgba(76, 175, 80, 0.2)',
                        fill: true,
                        tension: 0.3
                    }""" if snapshot_valid else "") + \
                ("," if snapshot_valid and ccs_valid else "") + \
                ("""
                    {
                        label: 'Commission Change (%)',
                        data: [""" + ", ".join(str(x) if x is not None else 'null' for x in commission_data) + """],
                        borderColor: '#800080',
                        backgroundColor: 'rgba(128, 0, 128, 0.2)',
                        fill: true,
                        tension: 0.3
                    }""" if commission_changes else "") + """
                ]
            },
            options: {
                responsive: true,
                maintainAspectRatio: false,
                scales: {
                    y: {
                        type: 'linear',
                        display: true,
                        title: {
                            display: true,
                            text: 'Percentage (%)'
                        },
                        beginAtZero: false
                    },
                    x: {
                        title: {
                            display: true,
                            text: 'Date'
                        }
                    }
                },
                plugins: {
                    legend: {
                        display: true,
                        position: 'top'
                    },
                    tooltip: {
                        callbacks: {
                            label: function(context) {
                                let label = context.dataset.label || '';
                                if (label) {
                                    label += ': ';
                                }
                                if (context.parsed.y !== null) {
                                    label += context.parsed.y + '%';
                                }
                                return label;
                            }
                        }
                    }
                }
            }
        });
    </script>
</body>
</html>
"""

        # Output to Reports subfolder in Program Data
        output_dir = os.path.join(data_dir, 'Reports')
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        html_path = os.path.join(output_dir, 'analytics.html')
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        webbrowser.open('file://' + os.path.realpath(html_path))

    except Exception as e:
        print(f"Error generating analytics: {str(e)}")
    finally:
        for file, temp_file in temp_files:
            original_path = os.path.join(data_dir, file)
            try:
                if os.path.exists(temp_file):
                    os.rename(temp_file, original_path)
            except Exception as e:
                print(f"Error restoring filename for {file}: {str(e)}")
                
# --- GENERATE CCS -------------------------------------------------------------------------------------------------

# UHC: Jarvis -> Commissions(Statements and More) -> Download
# Humana: AgentPortal -> Commission Statement Portal -> Excel

def process_aetna_cs(aetna_file, snapshot_file, permission_errors, os_errors, valid_generation):
    # 1) Read the Aetna file
    try:
        ext = os.path.splitext(aetna_file)[1].lower()
        if ext == '.csv':
            df = pd.read_csv(aetna_file, header=0, dtype=str)
        elif ext in ('.xls', '.xlsx'):
            df = pd.read_excel(aetna_file, header=0, dtype=str)
        else:
            valid_generation['state'] = False
            return
    except PermissionError:
        permission_errors['state'] = True
        valid_generation['state'] = False
        return
    except OSError:
        os_errors['state'] = True
        valid_generation['state'] = False
        return

    # Normalize column names
    df.columns = df.columns.str.strip()

    # 2) Define mapping and check presence
    mapping = {
        'Customer name': 'name',
        'Product 1': 'product',
        'Net compensation': 'YTD paid'
    }
    missing = [col for col in mapping if col not in df.columns]
    if missing:
        valid_generation['state'] = False
        return

    # 3) Subset and rename
    new_rows = df[list(mapping)].rename(columns=mapping)

    # 4) Read existing snapshot
    try:
        existing = pd.read_excel(snapshot_file, header=0, dtype=str)
    except PermissionError:
        permission_errors['state'] = True
        valid_generation['state'] = False
        return
    except OSError:
        os_errors['state'] = True
        valid_generation['state'] = False
        return

    # 5) Append and write back
    combined = pd.concat([existing, new_rows], ignore_index=True)
    try:
        with pd.ExcelWriter(snapshot_file, engine='openpyxl', mode='w') as writer:
            combined.to_excel(writer, index=False, sheet_name='Sheet1')
        valid_generation['state'] = True
    except PermissionError:
        permission_errors['state'] = True
        valid_generation['state'] = False
    except OSError:
        os_errors['state'] = True
        valid_generation['state'] = False

def process_humana_cs(humana_file, snapshot_file, permission_errors, os_errors, valid_generation):
    # 1) Read the Humana file
    try:
        ext = os.path.splitext(humana_file)[1].lower()
        if ext == '.csv':
            df = pd.read_csv(humana_file, header=0, dtype=str)
        elif ext in ('.xls', '.xlsx'):
            df = pd.read_excel(humana_file, header=0, dtype=str)
        else:
            print(f"[DEBUG] Unsupported Humana file ext: {ext}")
            valid_generation['state'] = False
            return
    except Exception as e:
        # print(f"[DEBUG] Error reading Humana file: {e}")
        permission_errors['state'] = isinstance(e, PermissionError)
        os_errors['state'] = isinstance(e, OSError)
        valid_generation['state'] = False
        return

    # Normalize and map columns
    df.columns = df.columns.str.strip()
    mapping = {'GrpName': 'name', 'Product': 'product', 'Comment': 'comment', 'PaidAmount': 'paid'}
    # Check presence
    missing_src = [c for c in mapping if c not in df.columns]
    if missing_src:
        # print(f"[DEBUG] Missing Humana columns: {missing_src}")
        valid_generation['state'] = False
        return

    # Subset & rename
    new_rows = df[list(mapping)].rename(columns=mapping)
    # print(f"[DEBUG] Humana new_rows shape: {new_rows.shape}")

    # 2) Read existing snapshot
    try:
        existing = pd.read_excel(snapshot_file, sheet_name=0, header=0, dtype=str)
        # print(f"[DEBUG] Existing snapshot shape: {existing.shape}")
    except Exception as e:
        # print(f"[DEBUG] Error reading snapshot: {e}")
        permission_errors['state'] = isinstance(e, PermissionError)
        os_errors['state'] = isinstance(e, OSError)
        valid_generation['state'] = False
        return

    # 3) Combine
    combined = pd.concat([existing, new_rows], ignore_index=True)
    # print(f"[DEBUG] Combined shape: {combined.shape}")

    # 4) Write back, explicitly to default sheet
    try:
        with pd.ExcelWriter(snapshot_file, engine='openpyxl', mode='w') as writer:
            combined.to_excel(writer, index=False, sheet_name='Sheet1')
        valid_generation['state'] = True
        # print("[DEBUG] Write successful")
    except Exception as e:
        # print(f"[DEBUG] Error writing snapshot: {e}")
        permission_errors['state'] = isinstance(e, PermissionError)
        os_errors['state'] = isinstance(e, OSError)
        valid_generation['state'] = False

def process_uhc_cs(uhc_file, snapshot_file, permission_errors, os_errors, valid_generation):
    # Attempt read
    try:
        ext = os.path.splitext(uhc_file)[1].lower()
        if ext == '.csv':
            df = pd.read_csv(uhc_file, header=0, dtype=str)
        elif ext in ('.xls', '.xlsx'):
            df = pd.read_excel(uhc_file, header=0, dtype=str)
        else:
            valid_generation['state'] = False
            return
    except PermissionError:
        permission_errors['state'] = True
        valid_generation['state'] = False
        return
    except OSError:
        os_errors['state'] = True
        valid_generation['state'] = False
        return

    # Normalize column names by stripping whitespace
    df.columns = df.columns.str.strip()
    df = df[df['Commission'].fillna('').str.strip() != '']

    # Define mapping from source to target
    mapping = {
        'Member Name': 'name',
        'Plan Type': 'product',
        'Commission Action': 'comment',
        'Commission': 'paid'
    }

    # Check for presence of all required source columns
    missing = [col for col in mapping if col not in df.columns]
    if missing:
        # Provide debug info in valid_generation if desired
        valid_generation['state'] = False
        return

    # Subset and rename
    output_df = df[list(mapping.keys())].rename(columns=mapping)

    # Try write
    try:
        ext_out = os.path.splitext(snapshot_file)[1].lower()
        output_df.to_excel(snapshot_file, index=False)
    except PermissionError:
        permission_errors['state'] = True
        valid_generation['state'] = False
    except OSError:
        os_errors['state'] = True
        valid_generation['state'] = False

    return output_df

def generate_ccs():
    valid_generation = {'state': True}
    permission_errors = {'state': False}
    os_errors = {'state': False}
    directories = [
        {'name': 'Reports',
         'found': False,
         'location': None},
        {'name': 'Snapshots',
         'found': False,
         'location': None},
        {'name': 'Program Data',
         'found': False,
         'location': None}
        ]
    carrier_reports = [
        {'name': 'UHC',
         'headers': ['UAD Activity', 'Reduction Indicator'],
         'found': False,
         'location': None},
        {'name': 'Humana',
         'headers': ['GrpName', 'PltfrmCd'],
         'found': False,
         'location': None},
        {'name': 'Aetna',
         'headers': ['Product 1', 'Product 2', 'Product 3', 'Product 4'],
         'found': False,
         'location': None}
        ]
    switch = 'CS'

    print("Generating CCS. This may take awhile...")

    print("Identifying directories...")
    find_reports_directory(directories, valid_generation, permission_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    elif not valid_generation['state']:
        print("Reports folder is empty. Cannot proceed with Snapshot generation.\n"
              + "Please populate the Reports folder with the necessary spreadsheet files and try again.")
        return
    reports_directory = directories[0]
    print("Reports directory found!")
    find_snapshots_directory(directories, permission_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    snapshots_directory = directories[1]
    print("Snapshots directory found!")
    find_data_directory(directories, permission_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    data_directory = directories[2]['location']
    print("Program Data directory found!")

    print("Identifying data...")
    convert_all_to_xlsx(reports_directory)
    find_uhc_data(reports_directory, carrier_reports, permission_errors, os_errors, switch)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    if os_errors['state']:
        print("Cannot access the UHC spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        return
    find_humana_data(reports_directory, carrier_reports, permission_errors, os_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    if os_errors['state']:
        print("Cannot access the Humana spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        return
    find_aetna_data(reports_directory, carrier_reports, permission_errors, os_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    if os_errors['state']:
        print("Cannot access the Aetna spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        return
    for carrier in carrier_reports:
        if not carrier['found']:
            valid_generation['state'] = False
            break
    if not valid_generation['state']:
        missing_reports = []
        for carrier in carrier_reports:
            if not carrier['found']:
                missing_reports.append(carrier['name'])
        print("Reports are missing for the following carriers:")
        for name in missing_reports:
            print(name)
        print("\nPlease populate the Reports folder with the appropriate spreadsheets and try again.")
        return
    print("Data found!")

    print("Processing data...")
    snapshots_directory = directories[1]
    print("Creating temporary data point...")
    snapshot_file = create_dated_spreadsheet(snapshots_directory, permission_errors, os_errors, switch)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    if os_errors['state']:
        print("Cannot create temporary data point in the Snapshots folder. Cannot proceed with Snapshot generation.")
    uhc_file = carrier_reports[0]['location']
    print("Processing UHC spreadsheet...")
    process_uhc_cs(uhc_file, snapshot_file, permission_errors, os_errors, valid_generation)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        return
    if os_errors['state']:
        print("Cannot access the UHC spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        if permission_errors['state']:
            print("Permission error encounter.\nPlease ensure that the file is not open in Excel and try again.")
        return
    print("Processing Humana spreadsheet...")
    humana_file = carrier_reports[1]['location']
    process_humana_cs(humana_file, snapshot_file, permission_errors, os_errors, valid_generation)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        return
    if os_errors['state']:
        print("Cannot access the UHC spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        if permission_errors['state']:
            print("Permission error encounter.\nPlease ensure that the file is not open in Excel and try again.")
        return
    print("Processing Aetna spreadsheet...")
    aetna_file = carrier_reports[2]['location']
    process_aetna_cs(aetna_file, snapshot_file, permission_errors, os_errors, valid_generation)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        return
    if os_errors['state']:
        print("Cannot access the UHC spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        if permission_errors['state']:
            print("Permission error encounter.\nPlease ensure that the file is not open in Excel and try again.")
        return
    copy_snapshot_to_data_directory(snapshot_file, data_directory, os_errors)
    print("Success!")

# --- GENERATE SNAPSHOT --------------------------------------------------------------------------------------------

# UHC: Jarvis -> Book of Business -> Download results
# Humana: AgentPortal -> View all Customers -> Export(Default settings) -> Export
# Devoted: Portal -> Clients -> Download CSV
# Aetna: Producer World -> Customer & Compensation Summary -> Book of business for: Integrity Insurance, LLC -> Download as CSV

def merge_base(s):
    non_empty = [v for v in s.dropna().astype(str) if v.strip()]
    if not non_empty: return ''
    uniq = list(set(non_empty))
    return uniq[0] if len(uniq)==1 else random.choice(uniq)

def pad_date(s):
    if pd.isna(s):
        return ''
    s = str(s).strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        mo, da, yr = m.groups()
        return f"{mo.zfill(2)}/{da.zfill(2)}/{yr}"
    return s

def process_aetna_spreadsheet(aetna_file, snapshot_file, permission_errors, os_errors):
    # Load
    try:
        _, ext = os.path.splitext(aetna_file.lower())
        if ext in ('.xls', '.xlsx'):
            df = pd.read_excel(aetna_file, header=0, dtype=str)
        elif ext == '.csv':
            df = pd.read_csv(aetna_file, dtype=str)
        else:
            raise ValueError(f"Unsupported extension: {ext!r}")
    except PermissionError:
        permission_errors['state'] = True
        return
    except OSError:
        os_errors['state'] = True
        return

    # normalize and verify
    df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
    for col in ('customer_name','customer_no','effective_date'):
        if col not in df.columns:
            raise KeyError(f"Aetna file is missing required column {col!r}")

    # split name, pad date, map identifier, build pk
    df[['first_name','last_name']] = df['customer_name'].apply(
        lambda x: pd.Series((lambda parts: (parts[0], parts[-1]))(str(x).split()))
    )
    df['effective_date_1'] = df['effective_date'].apply(pad_date)
    df['identifier_1']    = df['customer_no'].fillna('').astype(str).str.strip()
    df['pk']              = df['first_name'].str.strip() + df['last_name'].str.strip()

    # keep only needed cols
    df_aet = df[['first_name','last_name','identifier_1','effective_date_1','pk']]

    # Snapshot & Workbook
    try:
        snap = pd.read_excel(snapshot_file, sheet_name='Sheet', dtype=str)
    except Exception as e:
        os_errors['state'] = True
        return
    snap.columns = snap.columns.str.strip()
    snap['pk'] = snap['first_name'].str.strip() + snap['last_name'].str.strip()

    wb = load_workbook(snapshot_file)
    ws = wb['Sheet']
    headers = [cell.value for cell in ws[1] if cell.value]
    prefixes = ('product','identifier','effective_date','term_date','carrier')

    # Merge rows with on-the-fly headers
    for _, row in df_aet.iterrows():
        # locate existing client
        idx = snap.index[
            (snap['first_name'].str.strip() == row['first_name'])
            & (snap['last_name'].str.strip() == row['last_name'])
        ]
        if not idx.empty:
            i = idx[0]
            taken = [
                int(c.split('_')[1]) for c in snap.columns
                if c.startswith('product_')
                and pd.notna(snap.at[i, c])
                and str(snap.at[i, c]).strip()
            ]
            slot = max(taken) + 1 if taken else 1
        else:
            # append new blank row
            i = len(snap)
            new = {c: '' for c in snap.columns}
            new.update({
                'first_name':  row['first_name'],
                'last_name':   row['last_name'],
                'mbi':         '',
                'dob':         '',
                'email':       '',
                'phone':       '',
                'address':     '',
                'address_2':   '',
                'city':        '',
                'county':      '',
                'state':       '',
                'postal_code': ''
            })
            snap = pd.concat([snap, pd.DataFrame([new])], ignore_index=True)
            slot = 1

        # dynamically add only the headers we need for this slot
        for p in prefixes:
            h = f"{p}_{slot}"
            if h not in headers:
                headers.append(h)
                ws.cell(row=1, column=len(headers), value=h)

        # write Aetna fields
        snap.at[i, f'identifier_{slot}']     = row['identifier_1']
        snap.at[i, f'effective_date_{slot}'] = row['effective_date_1']
        snap.at[i, f'carrier_{slot}']        = 'AET'
        # product_{slot} remains blank (no plan name)

    # save the updated headers
    wb.save(snapshot_file)

    # drop helper and overwrite sheet
    snap.drop(columns=['pk'], inplace=True)
    try:
        with pd.ExcelWriter(snapshot_file,
                             engine='openpyxl',
                             mode='a',
                             if_sheet_exists='replace') as writer:
            snap.to_excel(writer, sheet_name='Sheet', index=False)
    except PermissionError:
        permission_errors['state'] = True
    except OSError:
        os_errors['state'] = True

def process_devoted_spreadsheet(devoted_file, snapshot_file, permission_errors, os_errors):
    try:
        _, ext = os.path.splitext(devoted_file.lower())
        if ext in ('.xls', '.xlsx'):
            df = pd.read_excel(devoted_file, header=0, dtype=str)
        elif ext == '.csv':
            df = pd.read_csv(devoted_file, dtype=str)
        else:
            raise ValueError(f"Unsupported extension: {ext!r}")
    except PermissionError:
        permission_errors['state'] = True
        return
    except OSError:
        os_errors['state'] = True
        return

    df.columns = (
        df.columns
        .str.strip()
        .str.lower()
        )

    initial_count = len(df)
    # print("Devoted columns after normalization:", df.columns.tolist()) # DEBUG
    # if 'last_name' not in df.columns or 'birth_date' not in df.columns:
    #    raise KeyError(f'Expected keys not found: {df.columns.tolist()}')

    # Drop inactive or deceased
    mask = (
        df['aor_policy_status'].fillna('').str.strip().eq('INACTIVE_POLICY')
        | (
            df['death_date'].notna()
            & df['death_date'].astype(str).str.strip().ne('')
            )
        )
    df = df.loc[~mask].copy()

    # print(f"Rows before filter: {initial_count}, after filter: {len(df)}")
    # print("Sample death_date values:", df['death_date'].dropna().unique()[:5])

    # Force MM/DD/YYYY on date fields
    for c in ('birth_date', 'start_date', 'end_date'):
        if c in df.columns:
            df[c] = df[c].apply(pad_date)

    # Reformat phone: xxxxxxxxxx → xxx-xxx-xxxx
    if 'primary_phone' in df.columns:
        df['primary_phone'] = (
            df['primary_phone']
              .fillna('')
              .astype(str)
              .str.replace(r'[^0-9]', '', regex=True)
              .apply(lambda x: f"{x[:3]}-{x[3:6]}-{x[6:10]}" if len(x)==10 else x)
        )

    # consolidate dupes
    key         = ['last_name','birth_date']
    policy_cols = ['pbp_name','member_record_locator','start_date','end_date']
    base_cols   = [
        'first_name','last_name','address','address2','city','county',
        'email','primary_phone','birth_date','state','zip_code'
    ]

    # Count how many slots we’ll need
    cnts      = df.groupby(key).size()
    max_dupes = int(cnts.max()) if not cnts.empty else 1

    def consolidate(group):
        out = {}
        # Merge base fields
        for c in base_cols:
            out[c] = merge_base(group[c]) if c in group else ''
        # Spread each policy into its own slot
        group = group.reset_index(drop=True)
        for i, row in group.iterrows():
            for c in policy_cols:
                colname = c if i == 0 else f"{c}_{i+1}"
                out[colname] = row.get(c, '') or ''
        return pd.Series(out)

    consolidated = [consolidate(g) for _, g in df.groupby(key, sort=False)]
    df_devoted   = pd.DataFrame(consolidated)

    # print("Consolidated Devoted columns:", df_devoted.columns.tolist())
    
    df_devoted['primary_key'] = (
        df_devoted['last_name'].str.strip()
      + df_devoted['birth_date'].str.strip()
    )

    # Merge into snapshot
    try:
        df_snap = pd.read_excel(snapshot_file, sheet_name='Sheet', dtype=str)
    except Exception as e:
        os_errors['state'] = True
        return
    df_snap['primary_key'] = (
        df_snap['last_name'].str.strip()
      + df_snap['dob'].str.strip()
    )

    # How many policy slots already in snapshot?
    exist_ns = [
        int(re.search(r'_(\d+)$', c).group(1))
        for c in df_snap.columns
        if c.startswith('product_')
    ]
    max_exist = max(exist_ns) if exist_ns else 0

    # How many Devoted slots did we create?
    dev_ns = [1] + [
        int(re.search(r'_(\d+)$', c).group(1))
        for c in df_devoted.columns
        if c.startswith('pbp_name_')
    ]
    max_dev = max(dev_ns) if dev_ns else 1

    max_slot = max(max_exist, max_dev)

    # Expand headers in the live workbook
    wb = load_workbook(snapshot_file)
    ws = wb['Sheet']
    headers = [cell.value for cell in ws[1] if cell.value]
    prefixes = ('product','identifier','effective_date','term_date','carrier')
    for n in range(1, max_slot+1):
        for p in prefixes:
            h = f"{p}_{n}"
            if h not in headers:
                headers.append(h)
                ws.cell(row=1, column=len(headers), value=h)
    wb.save(snapshot_file)

    # Reload df_snap so pandas sees new columns
    df_snap = pd.read_excel(snapshot_file, sheet_name='Sheet', dtype=str)
    df_snap['primary_key'] = (
        df_snap['last_name'].str.strip()
      + df_snap['dob'].str.strip()
    )

    # Merge each Devoted record
    for _, dev in df_devoted.iterrows():
        match_idx = df_snap.index[df_snap['primary_key'] == dev['primary_key']]
        if not match_idx.empty:
            i = match_idx[0]
            # find next open slot
            taken = [
                int(c.split('_')[1])
                for c in df_snap.columns
                if c.startswith('product_')
                and pd.notna(df_snap.at[i, c])
                and str(df_snap.at[i, c]).strip()
            ]
            slot = max(taken) + 1 if taken else 1

            for j in range(1, max_dev+1):
                suf = '' if j == 1 else f"_{j}"
                prod = dev.get(f"pbp_name{suf}")
                if pd.isna(prod) or not str(prod).strip():
                    continue
                df_snap.at[i, f"product_{slot}"]        = prod
                df_snap.at[i, f"identifier_{slot}"]     = dev.get(f"member_record_locator{suf}", '')
                df_snap.at[i, f"effective_date_{slot}"] = dev.get(f"start_date{suf}", '')
                df_snap.at[i, f"term_date_{slot}"]      = dev.get(f"end_date{suf}", '')
                df_snap.at[i, f"carrier_{slot}"]        = 'DEV'
                slot += 1

        else:
            # entirely new client row
            new_row = {col: '' for col in df_snap.columns}
            # base fields to snapshot columns
            new_row.update({
                'first_name': dev.get('first_name',''),
                'last_name':  dev.get('last_name',''),
                'mbi':        '',
                'dob':        dev.get('birth_date',''),
                'email':      dev.get('email',''),
                'phone':      dev.get('primary_phone',''),
                'address':    dev.get('address',''),
                'address_2':  dev.get('address2',''),
                'city':       dev.get('city',''),
                'county':     dev.get('county',''),
                'state':      dev.get('state',''),
                'postal_code':dev.get('zip_code',''),
            })
            slot = 1
            for j in range(1, max_dev+1):
                suf = '' if j == 1 else f"_{j}"
                prod = dev.get(f"pbp_name{suf}")
                if pd.isna(prod) or not str(prod).strip():
                    continue
                new_row[f'product_{slot}']        = prod
                new_row[f'identifier_{slot}']     = dev.get(f'member_record_locator{suf}','')
                new_row[f'effective_date_{slot}'] = dev.get(f'start_date{suf}','')
                new_row[f'term_date_{slot}']      = dev.get(f'end_date{suf}','')
                new_row[f'carrier_{slot}']        = 'DEV'
                slot += 1
            df_snap = pd.concat([df_snap, pd.DataFrame([new_row])], ignore_index=True)

    # Clean up
    df_snap.drop(columns=['primary_key'], inplace=True)

    # Write back into “Sheet” (replacing it)
    try:
        with pd.ExcelWriter(snapshot_file,
                             engine='openpyxl',
                             mode='a',
                             if_sheet_exists='replace') as writer:
            df_snap.to_excel(writer, sheet_name='Sheet', index=False)
    except PermissionError:
        permission_errors['state'] = True
    except OSError:
        os_errors['state'] = True

    return

def process_humana_spreadsheet(humana_file, snapshot_file, permission_errors, os_errors):
    try:
        _, ext = os.path.splitext(humana_file.lower())
        if ext in ('.xls', '.xlsx'):
            df = pd.read_excel(humana_file, header=0, dtype=str)
        elif ext == '.csv':
            df = pd.read_csv(humana_file, dtype=str)
        else:
            raise OSError
    except PermissionError:
        permission_errors['state'] = True
        return
    except OSError:
        os_errors['state'] = True
        return

    # Debug: Check input DataFrame columns
    # print(f"Input DataFrame columns: {df.columns.tolist()}")  # Debug
    # if 'MbrLastName' not in df.columns:
    #    raise ValueError("Required column 'MbrLastName' is missing from the Humana spreadsheet.")

    # drop inactive/deceased
    mask = (
        (df['Status'] == 'Inactive Policy')
      | (df['Status Reason'] == 'Deceased')
      | (df['Status Description'] == 'Member is deceased')
    )
    df = df.loc[~mask].copy()

    # replace 'Unavailable' in Phone/Email
    for c in ('Phone','Email'):
        if c in df: df[c] = df[c].replace('Unavailable','')

    # force leading-zero dates in these cols
    for c in ('Birth Date','Inactive Date','Effective Date'):
        if c in df:
            df[c] = (pd.to_datetime(df[c], format='%m/%d/%Y', errors='coerce')
                       .dt.strftime('%m/%d/%Y')
                     )

    # consolidate dupes
    key = ['MbrLastName','Birth Date']
    policy_cols = ['SalesProduct','Humana ID','Effective Date','Inactive Date']
    base_cols   = ['MbrFirstName','MbrLastName','MbrMiddleInit','Phone','Email','Birth Date']

    # Debug: Verify key columns exist
    # for k in key:
    #    if k not in df.columns:
    #        raise ValueError(f"Required column '{k}' is missing from the Humana spreadsheet.")

    # figure max policies any client has
    cnts = df.groupby(key).size()
    max_dups = int(cnts.max()) if not cnts.empty else 1
    # print(f"Max duplicates (policies per client): {max_dups}")  # Debug

    def consolidate(group):
        # print(f"Group columns: {group.columns.tolist()}")  # Debug
        out = {}
        # Base columns: consolidate non-policy data
        for c in base_cols:
            out[c] = merge_base(group[c]) if c in group else ''
        # Policy columns: assign each policy to its specific slot
        group = group.reset_index(drop=True)
        for i in range(len(group)):
            for c in policy_cols:
                colname = c if i==0 else f'{c}_{i+1}'
                out[colname] = group[c].iloc[i] if c in group and i < len(group) else ''
        # print(f"Consolidate output columns: {list(out.keys())}")  # Debug
        return pd.Series(out)

    consolidated = []
    for _, group in df.groupby(key, sort = False):
        consolidated.append(consolidate(group))
    df_humana = pd.DataFrame(consolidated)
    # print(f"df_humana columns after consolidate: {df_humana.columns.tolist()}")  # Debug

    # Merge into snapshot
    # helper to build primary_key
    # if 'MbrLastName' not in df_humana.columns or 'Birth Date' not in df_humana.columns:
    #    raise ValueError("Consolidate failed to produce required columns: 'MbrLastName' or 'Birth Date' missing in df_humana.")
    df_humana['primary_key'] = df_humana['MbrLastName'].str.strip() + df_humana['Birth Date'].str.strip()

    # load existing sheet into pandas
    try:
        df_snap = pd.read_excel(snapshot_file, sheet_name='Sheet', dtype=str)
    except PermissionError:
        permission_errors['state'] = True
        return
    df_snap['primary_key'] = df_snap['last_name'].str.strip() + df_snap['dob'].str.strip()

    # figure how many slots already exist
    exist = [int(re.search(r'_(\d+)$', c).group(1))
             for c in df_snap.columns if c.startswith('product_')]
    max_exist = max(exist) if exist else 0

    # figure how many Humana slots we need
    hum_slots = [1] + [int(re.search(r'_(\d+)$', c).group(1))
                       for c in df_humana.columns if '_' in c and c.startswith('SalesProduct_')]
    max_humana = max(hum_slots) if hum_slots else 1

    max_slot = max(max_exist, max_humana)

    # ensure headers out to max_slot in the live workbook
    wb = load_workbook(snapshot_file)
    ws = wb['Sheet']
    cols = [cell.value for cell in ws[1] if cell.value]
    prefixes = ('product','identifier','effective_date','term_date','carrier')
    for n in range(1, max_slot+1):
        for p in prefixes:
            h = f"{p}_{n}"
            if h not in cols:
                cols.append(h)
                ws.cell(row=1, column=len(cols), value=h)
    wb.save(snapshot_file)

    # reload after header-expansion
    df_snap = pd.read_excel(snapshot_file, sheet_name='Sheet', dtype=str)
    df_snap['primary_key'] = df_snap['last_name'].str.strip() + df_snap['dob'].str.strip()

    # for each Humana row, merge its policies into the next open slots or create new row
    for _, hum in df_humana.iterrows():
        pk_idx = df_snap.index[df_snap['primary_key']==hum['primary_key']]
        if not pk_idx.empty:
            # Update existing row
            i = pk_idx[0]
            # find taken slots
            taken = [int(c.split('_')[1])
                     for c in df_snap.columns
                     if c.startswith('product_')
                     and pd.notna(df_snap.at[i,c])
                     and df_snap.at[i,c] != '']
            next_slot = max(taken)+1 if taken else 1

            for j in range(1, max_humana+1):
                suf = '' if j==1 else f"_{j}"
                prod = hum.get(f"SalesProduct{suf}", '')
                # skip if product is NaN or empty
                if pd.isna(prod) or str(prod).strip() == '':
                    continue
                # print(f"Assigning product_{next_slot} for primary_key {hum['primary_key']}: prod={prod}")  # Debug
                df_snap.at[i, f"product_{next_slot}"]        = prod
                df_snap.at[i, f"identifier_{next_slot}"]     = hum.get(f"Humana ID{suf}", '')
                df_snap.at[i, f"effective_date_{next_slot}"] = hum.get(f"Effective Date{suf}", '')
                df_snap.at[i, f"term_date_{next_slot}"]      = hum.get(f"Inactive Date{suf}", '')
                df_snap.at[i, f"carrier_{next_slot}"]        = 'HUM'
                next_slot += 1
        else:
            # Create new row for Humana client
            new_row = {
                'first_name': hum.get('MbrFirstName', ''),
                'last_name': hum.get('MbrLastName', ''),
                'mbi': '',
                'dob': hum.get('Birth Date', ''),
                'email': hum.get('Email', ''),
                'phone': hum.get('Phone', ''),
                'address': '',
                'address_2': '',
                'city': '',
                'county': '',
                'state': '',
                'postal_code': ''
            }
            for j in range(1, max_humana+1):
                suf = '' if j==1 else f"_{j}"
                prod = hum.get(f"SalesProduct{suf}", '')
                # skip if product is NaN or empty
                if pd.isna(prod) or str(prod).strip() == '':
                    continue
                # print(f"New row product_{j} for primary_key {hum['primary_key']}: prod={prod}")  # Debug
                new_row[f"product_{j}"] = prod
                new_row[f"identifier_{j}"] = hum.get(f"Humana ID{suf}", '')
                new_row[f"effective_date_{j}"] = hum.get(f"Effective Date{suf}", '')
                new_row[f"term_date_{j}"] = hum.get(f"Inactive Date{suf}", '')
                new_row[f"carrier_{j}"] = 'HUM'

            # Ensure all columns exist in new_row
            for col in df_snap.columns:
                if col not in new_row:
                    new_row[col] = ''
            df_snap = pd.concat([df_snap, pd.DataFrame([new_row])], ignore_index=True)

    df_snap.drop(columns=['primary_key'], inplace=True)

    # overwrite the Sheet tab with the updated DataFrame
    try:
        with pd.ExcelWriter(snapshot_file,
                             engine='openpyxl',
                             mode='a',
                             if_sheet_exists='replace') as writer:
            df_snap.to_excel(writer, sheet_name='Sheet', index=False)
    except PermissionError:
        permission_errors['state'] = True
    except OSError:
        os_errors['state'] = True

    return

def process_uhc_spreadsheet(uhc_file, snapshot_file, permission_errors, os_errors):
    # print(f'uhc_file = {uhc_file}') # DEBUG
    # print(f'snapshot_file = {snapshot_file}') # DEBUG
    # print(f'permission_errors = {permission_errors}') # DEBUG
    # print(f'os_errors = {os_errors}') # DEBUG
    try:
        _, ext = os.path.splitext(uhc_file)
        ext = ext.lower()
        if ext in ('.xls', '.xlsx'):
            df = pd.read_excel(uhc_file, skiprows=4)
        elif ext == '.csv':
            df = pd.read_csv(uhc_file, skiprows=4)
        else:
            raise ValueError(f"Unsupported extension: {ext}")
    except PermissionError:
        permission_errors['state'] = True
        return
    except OSError:
        os_errors['state'] = True
        return

    # print("Columns:", df.columns.tolist()) # DEBUG
    # print("Raw policyTermDate sample:", df['policyTermDate'].head(5).tolist()) # DEBUG
    
    # Step 1a: Store original policyTermDate strings
    df['policyTermDate_raw'] = df['policyTermDate'].astype(str).str.strip()
    
    # Step 1b: Parse policyTermDate as datetime, but keep '2300-01-01' as a special case
    df['policyTermDate'] = pd.to_datetime(
        df['policyTermDate'],
        format='%Y-%m-%d',
        errors='coerce'
    )
    # print("Parsed sample:", df['policyTermDate'].head(5).tolist()) # DEBUG
    # print("Dtype:", df['policyTermDate'].dtype) # DEBUG

    # Step 2: filter by agentName
    df = df[df.get('agentName', '').astype(str).str.strip() == "INTEGRITY INSURANCE LLC"]
    # print("Step 2: filter by agentName") # DEBUG
    # print(f'df = {df}') # DEBUG

    # Step 3: drop past terms using datetime comparison, keeping '2300-01-01'
    today = pd.Timestamp.now().normalize()
    df = df[
        (df['policyTermDate'] >= today) |  # Valid future dates
        (df['policyTermDate'].isna() & (df['policyTermDate_raw'] == '2300-01-01'))  # Special case
    ]
    # print("Step 3: drop past terms with datetime comparison") # DEBUG
    # print(f'df = {df}') # DEBUG

    # Step 4: primary_key
    df['primary_key'] = (
        df['memberLastName'].astype(str).str.strip()
        + df['dateOfBirth'].astype(str).str.strip()
    )

    # Step 5: count policies
    counts = df['primary_key'].value_counts()
    policies_per_client = int(counts.max()) if not counts.empty else 1

    # Step 6: extra UHC columns for duplicates
    for n in range(2, policies_per_client+1):
        df[f'memberNumber_{n}'] = pd.NA
        df[f'policyEffectiveDate_{n}'] = pd.NaT
        df[f'policyTermDate_{n}'] = pd.NA
        df[f'planName_{n}'] = pd.NA

    # Step 7: consolidate into one row per primary_key
    consolidated = []
    for _, group in df.groupby('primary_key', sort=False):
        group = group.reset_index(drop=True)
        base = group.loc[0].copy()
        for i in range(1, len(group)):
            n = i+1
            if pd.notna(group.loc[i, 'planName']):
                base[f"memberNumber_{n}"] = group.loc[i, 'memberNumber']
                base[f"policyEffectiveDate_{n}"] = group.loc[i, 'policyEffectiveDate']
                base[f"policyTermDate_{n}"] = group.loc[i, 'policyTermDate_raw']
                base[f"planName_{n}"] = group.loc[i, 'planName']
            else:
                base[f"memberNumber_{n}"] = pd.NA
                base[f"policyEffectiveDate_{n}"] = pd.NaT
                base[f"policyTermDate_{n}"] = pd.NA
                base[f"planName_{n}"] = pd.NA
        base['policyTermDate_1'] = base['policyTermDate_raw']
        consolidated.append(base)
    df = pd.DataFrame(consolidated)
    # print("Step 7: Consolidated DataFrame") # DEBUG
    # print(f'df shape: {df.shape}') # DEBUG
    # print(f'df columns: {df.columns.tolist()}') # DEBUG
    # print(f'df head: {df.head().to_string()}') # DEBUG
    # if df.empty:
        # print("WARNING: Consolidated DataFrame is empty. No data to writer.") # DEBUG

    wb = load_workbook(snapshot_file)
    ws = wb['Sheet'] if 'Sheet' in wb.sheetnames else wb.active

    # 8a) read whatever headers are already in row 1
    headers = [cell.value for cell in ws[1] if cell.value]

    # 8b) for each policy index (1 through N), make sure prefix_{i} is in the header row
    prefixes = ('identifier', 'effective_date', 'term_date', 'product', 'carrier')
    for i in range(1, policies_per_client + 1):
        for p in prefixes:
            h = f"{p}_{i}"
            if h not in headers:
                headers.append(h)
                ws.cell(row=1, column=len(headers), value=h)

    wb.save(snapshot_file)

    # Step 9: map and append to snapshot
    # 9a) map base columns into a fresh DataFrame
    base_map = {
        'memberFirstName': 'first_name',
        'memberLastName': 'last_name',
        'dateOfBirth': 'dob',
        'memberEmail': 'email',
        'memberPhone': 'phone',
        'memberAddress1': 'address',
        'memberAddress2': 'address_2',
        'memberCity': 'city',
        'memberState': 'state',
        'memberZip': 'postal_code',
        'memberCounty': 'county',
        'memberNumber': 'identifier',
        'policyEffectiveDate': 'effective_date',
        'policyTermDate_1': 'term_date',
        'mbiNumber': 'mbi',
        'planName': 'product'
    }
    snap_df = df[list(base_map)].rename(columns=base_map)

    # 9b) add carrier for policy #1
    snap_df['carrier'] = 'UHC'

    # 9c) add extra-policy columns (2..N), only if product exists
    for n in range(2, policies_per_client + 1):
        mask = pd.notna(df.get(f'planName_{n}', pd.NA))
        snap_df[f'identifier_{n}'] = df.get(f'memberNumber_{n}', pd.NA).where(mask, pd.NA)
        snap_df[f'effective_date_{n}'] = df.get(f'policyEffectiveDate_{n}', pd.NaT).where(mask, pd.NaT)
        snap_df[f'term_date_{n}'] = df.get(f'policyTermDate_{n}', pd.NA).where(mask, pd.NA)
        snap_df[f'product_{n}'] = df.get(f'planName_{n}', pd.NA).where(mask, pd.NA)
        # Fix: Create a Series with 'UHC' and apply mask
        snap_df[f'carrier_{n}'] = pd.Series('UHC', index=snap_df.index).where(mask, pd.NA)

    # 9d) rename policy #1 columns to include _1 so they match your sheet headers
    snap_df.rename(columns={
        'identifier': 'identifier_1',
        'effective_date': 'effective_date_1',
        'term_date': 'term_date_1',
        'product': 'product_1',
        'carrier': 'carrier_1'
    }, inplace=True)

    # 9e) read the live header row from the existing "Sheet" tab
    real_headers = pd.read_excel(
        snapshot_file,
        sheet_name='Sheet',
        nrows=0
    ).columns.tolist()

    # 9f) reindex to that exact column order
    snap_df = snap_df[real_headers]
    for col in snap_df.columns:
        if col.startswith('term_date_'):
            snap_df[col] = snap_df[col].astype(str).replace({'NaT': '', '<NA>': ''})

    # 9g) append beneath row 1 (headers already present)
    try:
        with pd.ExcelWriter(snapshot_file,
                            mode='a',
                            engine='openpyxl',
                            if_sheet_exists='overlay') as writer:
            snap_df.to_excel(
                writer,
                sheet_name='Sheet',
                index=False,
                header=False,
                startrow=1
            )
    except PermissionError:
        permission_errors['state'] = True
    except OSError:
        os_errors['state'] = True

    # Step 10: done
    return

def find_devoted_data(reports_directory, carrier_reports, permission_errors, os_errors):
    folder = reports_directory.get('location')
    devoted_entry = next((r for r in carrier_reports if r['name'] == 'Devoted'), None)
    expected_headers = devoted_entry['headers']

    try:
        for fname in os.listdir(folder):
            path = os.path.join(folder, fname)
            _, ext = os.path.splitext(fname.lower())
            if ext in ('.csv', '.xls', '.xlsx'):
                try:
                    if ext == '.csv':
                        df = pd.read_csv(path, header=0, dtype=str)
                    else:
                        df = pd.read_excel(path, header=0, dtype=str)
                except PermissionError:
                    # Cannot open/read this file
                    permission_errors['state'] = True
                    return

                # Check for exact header match
                cols = list(df.columns)
                if all(h in cols for h in expected_headers):
                    devoted_entry['found'] = True
                    devoted_entry['location'] = path
                    break

    except PermissionError:
        permission_errors['state'] = True
    except OSError:
        os_errors['state'] = True
    return

def generate_snapshot():
    valid_generation = {'state': True}
    permission_errors = {'state': False}
    os_errors = {'state': False}
    directories = [
        {'name': 'Reports',
         'found': False,
         'location': None},
        {'name': 'Snapshots',
         'found': False,
         'location': None},
        {'name': 'Program Data',
         'found': False,
         'location': None}
        ]
    carrier_reports = [
        {'name': 'UHC',
         'headers': ['memberFirstName', 'memberLastName', 'dateOfBirth', 'memberEmail'],
         'found': False,
         'location': None},
        {'name': 'Humana',
         'headers': ['MbrLastName', 'MbrFirstName', 'Humana ID', 'MbrMiddleInit'],
         'found': False,
         'location': None},
        {'name': 'Devoted',
         'headers': ['member_record_locator', 'pbp_name', 'birth_date', 'primary_phone'],
         'found': False,
         'location': None},
        {'name': 'Aetna',
         'headers': ['Product 1', 'Product 2', 'Product 3', 'Product 4'],
         'found': False,
         'location': None}
        ]
    switch = 'GS'

    print("Generating Snapshot. This may take awhile...")
    
    print("Identifying directories...")
    find_reports_directory(directories, valid_generation, permission_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    elif not valid_generation['state']:
        print("Reports folder is empty. Cannot proceed with Snapshot generation.\n"
              + "Please populate the Reports folder with the necessary spreadsheet files and try again.")
        return
    print("Reports directory found!")
    find_snapshots_directory(directories, permission_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    print("Snapshots directory found!")
    find_data_directory(directories, permission_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    print("Program Data directory found!")

    print("Identifying data...")
    reports_directory = directories[0]
    convert_all_to_xlsx(reports_directory)
    find_uhc_data(reports_directory, carrier_reports, permission_errors, os_errors, switch)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    if os_errors['state']:
        print("Cannot access the UHC spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        return
    find_humana_data(reports_directory, carrier_reports, permission_errors, os_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    if os_errors['state']:
        print("Cannot access the Humana spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        return
    find_devoted_data(reports_directory, carrier_reports, permission_errors, os_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    if os_errors['state']:
        print("Cannot access the Devoted spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        return
    find_aetna_data(reports_directory, carrier_reports, permission_errors, os_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    if os_errors['state']:
        print("Cannot access the Aetna spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        return
    for carrier in carrier_reports:
        if not carrier['found']:
            valid_generation['state'] = False
            break
    if not valid_generation['state']:
        missing_reports = []
        for carrier in carrier_reports:
            if not carrier['found']:
                missing_reports.append(carrier['name'])
        print("Reports are missing for the following carriers:")
        for name in missing_reports:
            print(name)
        print("\nPlease populate the Reports folder with the appropriate spreadsheets and try again.")
        return
    print("Data found!")

    print("Processing data...")
    snapshots_directory = directories[1]
    print("Creating temporary data point...")
    snapshot_file = create_dated_spreadsheet(snapshots_directory, permission_errors, os_errors, switch)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        return
    if os_errors['state']:
        print("Cannot create temporary data point in the Snapshots folder. Cannot proceed with Snapshot generation.")
    print("Processing UHC spreadsheet...")
    uhc_file = carrier_reports[0]['location']
    process_uhc_spreadsheet(uhc_file, snapshot_file, permission_errors, os_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        return
    if os_errors['state']:
        print("Cannot access the UHC spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        if permission_errors['state']:
            print("Permission error encounter.\nPlease ensure that the file is not open in Excel and try again.")
        return
    print("Processing Humana spreadsheet...")
    humana_file = carrier_reports[1]['location']
    process_humana_spreadsheet(humana_file, snapshot_file, permission_errors, os_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        return
    if os_errors['state']:
        print("Cannot access the UHC spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        if permission_errors['state']:
            print("Permission error encounter.\nPlease ensure that the file is not open in Excel and try again.")
        return
    print("Processing Devoted spreadsheet...")
    devoted_file = carrier_reports[2]['location']
    process_devoted_spreadsheet(devoted_file, snapshot_file, permission_errors, os_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        return
    if os_errors['state']:
        print("Cannot access the UHC spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        if permission_errors['state']:
            print("Permission error encounter.\nPlease ensure that the file is not open in Excel and try again.")
        return
    print("Processing Aetna spreadsheet...")
    aetna_file = carrier_reports[3]['location']
    process_aetna_spreadsheet(aetna_file, snapshot_file, permission_errors, os_errors)
    if permission_errors['state']:
        print("Permission error encounter. Cannot proceed with Snapshot generation.\n"
              + "Please re-launch the program with administrator privileges and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        return
    if os_errors['state']:
        print("Cannot access the UHC spreadsheet file. Cannot proceed with Snapshot generation.\n"
              + "Please ensure that the file is not open in Excel and try again.")
        delete_snapshot_file(snapshot_file, permission_errors, os_errors)
        if permission_errors['state']:
            print("Permission error encounter.\nPlease ensure that the file is not open in Excel and try again.")
        return
    print("Creating backup file...")
    data_directory = directories[2]['location']
    copy_snapshot_to_data_directory(snapshot_file, data_directory, os_errors)
    print("Success!")
